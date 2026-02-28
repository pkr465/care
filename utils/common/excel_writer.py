# excel_writer.py
"""
Standalone Excel Writer for codebase analysis reports.

Provides:
- Professional Excel workbook creation with openpyxl
- Metadata summary sheets with styled key-value tables
- Data table sheets with headers, borders, conditional formatting
- Auto-fit column widths, freeze panes, auto-filters
- Configurable pass/fail colors and labels
- Summary row generation (totals, counts)

Dependencies: openpyxl (pip install openpyxl)
Optional:     utils.parsers.env_parser.EnvConfig (for color config)

Usage:
    from utils.common.excel_writer import ExcelWriter

    writer = ExcelWriter("report.xlsx")
    writer.add_data_sheet({"Project": "App", "Date": "2025-01-01"}, "Info", "Analysis Report")
    writer.add_table_sheet(
        headers=["File", "Issues", "Status"],
        data_rows=[["main.py", 3, "FAIL"], ["utils.py", 0, "PASS"]],
        sheet_name="Results",
        status_column="Status",
    )
    writer.save()
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

# Lazy import openpyxl to provide clear error messages
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

@dataclass
class ExcelStyle:
    """Configurable style settings for Excel reports."""

    # Header styling
    header_font_color: str = "FFFFFF"
    header_bg_color: str = "4F81BD"
    header_bold: bool = True

    # Pass/Fail colors (hex without #)
    pass_color: str = "C6EFCE"
    fail_color: str = "FFC7CE"
    warn_color: str = "FFEB9C"
    alt_row_color: str = "F3F3F3"

    # Pass/Fail labels (case-insensitive matching)
    pass_labels: Set[str] = field(default_factory=lambda: {
        "PASS", "PASSED", "OK", "SUCCESS", "YES", "TRUE", "CLEAN", "FIXED",
    })
    fail_labels: Set[str] = field(default_factory=lambda: {
        "FAIL", "FAILED", "ERROR", "NO", "FALSE", "CRITICAL", "BUG",
    })
    warn_labels: Set[str] = field(default_factory=lambda: {
        "WARN", "WARNING", "SKIP", "SKIPPED", "PENDING", "PARTIAL",
    })

    # Status column detection (case-insensitive)
    status_column_names: Set[str] = field(default_factory=lambda: {
        "status", "result", "pass/fail", "outcome", "verdict",
    })

    # Layout
    min_column_width: int = 12
    max_column_width: int = 60
    freeze_header: bool = True
    auto_filter: bool = True

    @classmethod
    def from_env(cls, env_config=None) -> "ExcelStyle":
        """Build style config from EnvConfig or environment variables."""
        import os

        if env_config is None:
            try:
                from utils.parsers.env_parser import EnvConfig
                env_config = EnvConfig()
            except ImportError:
                env_config = None

        def _get(key: str, default: str = "") -> str:
            if env_config and hasattr(env_config, "get"):
                val = env_config.get(key)
                if val is not None and val != "":
                    return str(val)
            return os.getenv(key, default)

        return cls(
            pass_color=_get("PASS_COLOR", "C6EFCE"),
            fail_color=_get("FAIL_COLOR", "FFC7CE"),
            alt_row_color=_get("ALT_ROW_COLOR", "F3F3F3"),
        )


# ---------------------------------------------------------------------------
# Excel Writer
# ---------------------------------------------------------------------------

class ExcelWriter:
    """
    Professional Excel report writer for codebase analysis output.

    Features:
    - Metadata sheets with styled key-value tables
    - Data table sheets with formatted headers and conditional formatting
    - Auto-detected pass/fail/warn status coloring
    - Auto-fit column widths with min/max constraints
    - Freeze panes on header row
    - Auto-filter on data columns
    - Summary row generation
    - Configurable styles via ExcelStyle or EnvConfig

    Usage:
        writer = ExcelWriter("report.xlsx")
        writer.add_data_sheet(metadata_dict, "Metadata", "Report Title")
        writer.add_table_sheet(headers, rows, "Results")
        writer.save()
    """

    def __init__(
        self,
        file_path: Union[str, Path],
        style: Optional[ExcelStyle] = None,
        env_config=None,
    ):
        """
        Initialize a new Excel workbook.

        Args:
            file_path: Destination path for the generated Excel file.
            style: Optional ExcelStyle configuration.
            env_config: Optional EnvConfig for style defaults.
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "The 'openpyxl' package is required. "
                "Install it with: pip install openpyxl"
            )

        self.file_path = Path(file_path)
        self.style = style or ExcelStyle.from_env(env_config)
        self.wb = openpyxl.Workbook()
        self._remove_default_sheet()
        self.logger = logger

    def _remove_default_sheet(self) -> None:
        """Remove the default 'Sheet' created by openpyxl."""
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    # ------------------------------------------------------------------
    # Metadata Sheet
    # ------------------------------------------------------------------

    def add_data_sheet(
        self,
        data_dict: Dict[str, Any],
        title: str = "Metadata",
        report_title: str = "Analysis Report",
    ) -> Worksheet:
        """
        Add a metadata summary sheet with styled key-value pairs.

        Args:
            data_dict: Mapping of parameter names to values.
            title: Worksheet tab name.
            report_title: Main title displayed at the top.

        Returns:
            The created Worksheet.
        """
        ws = self.wb.create_sheet(title)

        # Title row
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = report_title
        title_cell.font = Font(size=16, bold=True, color="2C3E50")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column headers
        for col_letter, header_text in [("A", "Parameter"), ("B", "Value")]:
            cell = ws[f"{col_letter}2"]
            cell.value = header_text
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 65

        # Data rows
        thin_border = Border(
            bottom=Side(border_style="thin", color="DDDDDD"),
        )
        for i, (key, value) in enumerate(data_dict.items(), start=3):
            key_cell = ws[f"A{i}"]
            val_cell = ws[f"B{i}"]

            key_cell.value = key
            key_cell.font = Font(bold=True, color="333333")
            key_cell.fill = PatternFill("solid", fgColor="F5F5F5")
            key_cell.alignment = Alignment(horizontal="right", vertical="top")
            key_cell.border = thin_border

            val_cell.value = str(value) if value is not None else ""
            val_cell.alignment = Alignment(wrap_text=True, vertical="top")
            val_cell.border = thin_border

        # Freeze header
        ws.freeze_panes = "A3"

        return ws

    # ------------------------------------------------------------------
    # Data Table Sheet
    # ------------------------------------------------------------------

    def add_table_sheet(
        self,
        headers: List[str],
        data_rows: List[Union[List, Dict]],
        sheet_name: str = "Results",
        status_column: Optional[str] = None,
        conditional_formats: Optional[Callable] = None,
        autofit: bool = True,
        add_summary: bool = False,
    ) -> Worksheet:
        """
        Add a data table sheet with formatted headers and optional conditional formatting.

        Args:
            headers: Column header names.
            data_rows: Row data as lists or dicts (keyed by header names).
            sheet_name: Worksheet tab name.
            status_column: Column name to apply pass/fail coloring.
                If None, auto-detects from style.status_column_names.
            conditional_formats: Custom callback(cell, col_name, cell_value, row_idx).
            autofit: Auto-adjust column widths to fit content.
            add_summary: Add a summary row at the bottom.

        Returns:
            The created Worksheet.
        """
        ws = self.wb.create_sheet(sheet_name)

        # --- Write headers ---
        header_font = Font(bold=True, color=self.style.header_font_color)
        header_fill = PatternFill("solid", fgColor=self.style.header_bg_color)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # --- Detect status column ---
        status_col_idx = None
        if status_column:
            for idx, h in enumerate(headers):
                if h.strip().lower() == status_column.strip().lower():
                    status_col_idx = idx
                    break
        else:
            for idx, h in enumerate(headers):
                if h.strip().lower() in self.style.status_column_names:
                    status_col_idx = idx
                    break

        # --- Write data rows ---
        thin = Side(border_style="thin", color="D0D0D0")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        pass_fill = PatternFill("solid", fgColor=self.style.pass_color)
        fail_fill = PatternFill("solid", fgColor=self.style.fail_color)
        warn_fill = PatternFill("solid", fgColor=self.style.warn_color)
        alt_fill = PatternFill("solid", fgColor=self.style.alt_row_color)

        for row_idx, row in enumerate(data_rows, 2):
            # Normalize row to list
            if isinstance(row, dict):
                row_values = [row.get(h, "") for h in headers]
            else:
                row_values = list(row)

            # Pad or trim to header count
            row_values = row_values[:len(headers)]
            while len(row_values) < len(headers):
                row_values.append("")

            for col_idx, val in enumerate(row_values, 1):
                # Sanitize: openpyxl only accepts str, int, float, bool,
                # datetime, or None.  Convert anything else to str.
                if val is not None and not isinstance(val, (str, int, float, bool)):
                    try:
                        # datetime-like objects are fine
                        import datetime
                        if not isinstance(val, (datetime.datetime, datetime.date, datetime.time)):
                            val = str(val)
                    except Exception:
                        val = str(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = cell_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Custom conditional formatting callback
                if conditional_formats:
                    col_name = headers[col_idx - 1].strip().lower()
                    cell_str = str(val).strip().upper()
                    conditional_formats(cell, col_name, cell_str, row_idx)
                    continue

                # Auto status coloring
                if status_col_idx is not None and (col_idx - 1) == status_col_idx:
                    upper_val = str(val).strip().upper()
                    if upper_val in self.style.pass_labels:
                        cell.fill = pass_fill
                        cell.font = Font(bold=True, color="006100")
                    elif upper_val in self.style.fail_labels:
                        cell.fill = fail_fill
                        cell.font = Font(bold=True, color="9C0006")
                    elif upper_val in self.style.warn_labels:
                        cell.fill = warn_fill
                        cell.font = Font(bold=True, color="9C6500")
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill

        # --- Auto-fit columns ---
        if autofit:
            self._autofit_columns(ws, headers, len(data_rows) + 1)

        # --- Freeze header row ---
        if self.style.freeze_header:
            ws.freeze_panes = "A2"

        # --- Auto-filter ---
        if self.style.auto_filter and data_rows:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{len(data_rows) + 1}"

        # --- Summary row ---
        if add_summary:
            self._add_summary_row(ws, headers, data_rows)

        return ws

    # ------------------------------------------------------------------
    # Summary Row
    # ------------------------------------------------------------------

    def _add_summary_row(
        self,
        ws: Worksheet,
        headers: List[str],
        data_rows: List,
    ) -> None:
        """Add a summary/totals row at the bottom of the data table."""
        summary_row = len(data_rows) + 2  # +1 for header, +1 for next row
        summary_fill = PatternFill("solid", fgColor="D9E1F2")
        summary_font = Font(bold=True, color="1F4E79")
        summary_border = Border(
            top=Side(border_style="medium", color="4F81BD"),
            bottom=Side(border_style="medium", color="4F81BD"),
        )

        ws.cell(row=summary_row, column=1, value="SUMMARY")
        ws.cell(row=summary_row, column=1).font = summary_font
        ws.cell(row=summary_row, column=1).fill = summary_fill
        ws.cell(row=summary_row, column=1).border = summary_border

        for col_idx, header in enumerate(headers[1:], 2):
            cell = ws.cell(row=summary_row, column=col_idx)
            cell.fill = summary_fill
            cell.font = summary_font
            cell.border = summary_border

            # Try to compute a sum for numeric columns
            values = []
            for row in data_rows:
                if isinstance(row, dict):
                    v = row.get(header, "")
                elif isinstance(row, (list, tuple)) and col_idx - 1 < len(row):
                    v = row[col_idx - 1]
                else:
                    v = ""
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    pass

            if values:
                cell.value = sum(values)
            else:
                # Count non-empty values
                count = sum(
                    1 for row in data_rows
                    for v in ([row.get(header, "")] if isinstance(row, dict) else [row[col_idx - 1] if col_idx - 1 < len(row) else ""])
                    if str(v).strip()
                )
                cell.value = f"{count} entries"

    # ------------------------------------------------------------------
    # Column Auto-fit
    # ------------------------------------------------------------------

    def _autofit_columns(
        self, ws: Worksheet, headers: List[str], last_row: int
    ) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, last_row + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    # Handle multi-line cells
                    lines = str(cell_val).split("\n")
                    longest = max(len(line) for line in lines)
                    max_len = max(max_len, longest)

            # Apply constraints
            width = min(
                max(max_len + 3, self.style.min_column_width),
                self.style.max_column_width,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ------------------------------------------------------------------
    # Convenience Methods
    # ------------------------------------------------------------------

    def is_pass(self, value: Any) -> bool:
        """Check if a value matches a 'pass' label."""
        return str(value).strip().upper() in self.style.pass_labels

    def is_fail(self, value: Any) -> bool:
        """Check if a value matches a 'fail' label."""
        return str(value).strip().upper() in self.style.fail_labels

    def is_warn(self, value: Any) -> bool:
        """Check if a value matches a 'warn' label."""
        return str(value).strip().upper() in self.style.warn_labels

    def add_sheet_from_dicts(
        self,
        records: List[Dict[str, Any]],
        sheet_name: str = "Data",
        status_column: Optional[str] = None,
    ) -> Worksheet:
        """
        Convenience: create a table sheet from a list of dicts.
        Automatically extracts headers from dict keys.

        Args:
            records: List of dicts (each dict is a row).
            sheet_name: Worksheet tab name.
            status_column: Column for pass/fail coloring.

        Returns:
            The created Worksheet.
        """
        if not records:
            self.logger.warning("No records to write for sheet '%s'", sheet_name)
            headers = ["(empty)"]
            data_rows = []
        else:
            # Collect all unique keys preserving order
            seen = set()
            headers = []
            for rec in records:
                for key in rec.keys():
                    if key not in seen:
                        seen.add(key)
                        headers.append(key)
            data_rows = records

        return self.add_table_sheet(
            headers=headers,
            data_rows=data_rows,
            sheet_name=sheet_name,
            status_column=status_column,
        )

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------

    def save(self) -> Path:
        """
        Save the workbook to disk.

        Creates parent directories if they don't exist.

        Returns:
            The Path where the file was saved.
        """
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self.file_path))
        self.logger.info("Excel report saved: %s", self.file_path)
        return self.file_path

    def __repr__(self) -> str:
        sheets = self.wb.sheetnames
        return f"ExcelWriter(path='{self.file_path}', sheets={sheets})"
