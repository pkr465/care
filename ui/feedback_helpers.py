"""
feedback_helpers.py

CARE — Codebase Analysis & Repair Engine
Helpers for converting between DataFrames, JSONL directives, Excel reports,
and QA traceability reports.

Author: Pavan R
"""

import io
import json
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
#  Results → DataFrame
# ═══════════════════════════════════════════════════════════════════════════════

def results_to_dataframe(results: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Convert CodebaseLLMAgent analysis results to a DataFrame
    with added editable feedback columns.

    Args:
        results: List of issue dicts from the agent, each containing:
            File, Title, Severity, Confidence, Category, Line,
            Description, Suggestion, Code, Fixed_Code

    Returns:
        DataFrame with original columns plus Action and Notes columns.
    """
    if not results:
        return pd.DataFrame(columns=[
            "File", "Title", "Severity", "Confidence", "Category",
            "Line", "Description", "Suggestion", "Code", "Fixed_Code",
            "Action", "Notes",
        ])

    df = pd.DataFrame(results)

    # Ensure expected columns exist
    expected = [
        "File", "Title", "Severity", "Confidence", "Category",
        "Line", "Description", "Suggestion", "Code", "Fixed_Code",
    ]
    for col in expected:
        if col not in df.columns:
            df[col] = ""

    # Add editable feedback columns
    if "Action" not in df.columns:
        df["Action"] = "Auto-fix"
    if "Notes" not in df.columns:
        df["Notes"] = ""

    # Reorder columns: feedback columns at the end
    ordered = [c for c in expected if c in df.columns] + ["Action", "Notes"]
    extra = [c for c in df.columns if c not in ordered]
    df = df[ordered + extra]

    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  DataFrame → JSONL Directives
# ═══════════════════════════════════════════════════════════════════════════════

ACTION_MAP = {
    "Auto-fix": "FIX",
    "Skip": "SKIP",
    "Review": "FIX_WITH_CONSTRAINTS",
    "Manual Review": "FIX_WITH_CONSTRAINTS",
}


def dataframe_to_directives(
    df: pd.DataFrame,
    output_path: str,
) -> str:
    """
    Convert an edited feedback DataFrame to JSONL directives
    for CodebaseFixerAgent.

    Args:
        df: DataFrame with columns including File, Line, Category,
            Severity, Suggestion, Fixed_Code, Action, Notes
        output_path: Path to write the JSONL file

    Returns:
        Path to the written JSONL file.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    directives = []
    for _, row in df.iterrows():
        action_label = str(row.get("Action", "Auto-fix")).strip()
        action = ACTION_MAP.get(action_label, "FIX")

        # Skip issues explicitly marked as Skip
        if action == "SKIP":
            continue

        directive = {
            "file_path": str(row.get("File", "")),
            "line_number": int(row.get("Line", 0)) if pd.notna(row.get("Line")) else 0,
            "issue_type": str(row.get("Category", "")),
            "severity": str(row.get("Severity", "Medium")),
            "title": str(row.get("Title", "")),
            "description": str(row.get("Description", "")),
            "suggested_fix": str(row.get("Suggestion", "")),
            "fixed_code": str(row.get("Fixed_Code", "")),
            "action": action,
            "source_type": "llm",
        }

        # Include user notes as constraints if provided
        notes = str(row.get("Notes", "")).strip()
        if notes:
            directive["constraints"] = notes

        directives.append(directive)

    with open(output_path, "w", encoding="utf-8") as f:
        for d in directives:
            f.write(json.dumps(d, default=str) + "\n")

    logger.info(
        "Wrote %d directives to %s (%d skipped)",
        len(directives),
        output_path,
        len(df) - len(directives),
    )
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  Save feedback back to detailed_code_review.xlsx
# ═══════════════════════════════════════════════════════════════════════════════

def save_feedback_to_excel(
    df: pd.DataFrame,
    excel_path: str,
) -> bool:
    """
    Write Action/Notes feedback from the Review UI back into the
    existing detailed_code_review.xlsx file.

    Maps:
        Action  → "Feedback" column in the Excel
        Notes   → "Constraints" column in the Excel

    If the Excel file does not exist, creates a new one.

    Returns True on success, False on failure.
    """
    try:
        import openpyxl

        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
        else:
            # Create a fresh workbook with the Analysis sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis"
            headers = [
                "S.No", "Title", "Severity", "Confidence", "Category",
                "File", "Line", "Description", "Suggestion",
                "Code", "Fixed_Code", "Feedback", "Constraints",
            ]
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=h)
            # Populate rows from DataFrame
            for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                for col_idx, col_name in enumerate(headers[1:], 2):
                    mapped = col_name
                    if col_name == "Feedback":
                        mapped = "Action"
                    elif col_name == "Constraints":
                        mapped = "Notes"
                    val = row.get(mapped, "")
                    if val is not None and not isinstance(val, (str, int, float, bool)):
                        val = str(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(excel_path)
            logger.info("Created new Excel with feedback: %s", excel_path)
            return True

        # --- Update existing workbook ---
        ws = None
        for name in wb.sheetnames:
            if name.lower() in ("analysis", "sheet1"):
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Build header → column index map (1-based)
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col_idx).value
            if cell_val:
                header_map[str(cell_val).strip()] = col_idx

        # Ensure Feedback and Constraints columns exist
        next_col = ws.max_column + 1
        if "Feedback" not in header_map:
            ws.cell(row=1, column=next_col, value="Feedback")
            header_map["Feedback"] = next_col
            next_col += 1
        if "Constraints" not in header_map:
            ws.cell(row=1, column=next_col, value="Constraints")
            header_map["Constraints"] = next_col

        feedback_col = header_map["Feedback"]
        constraints_col = header_map["Constraints"]

        # Build a lookup from (File, Line, Title) → (Action, Notes)
        feedback_lookup = {}
        for _, row in df.iterrows():
            key = (
                str(row.get("File", "")).strip(),
                str(row.get("Line", "")).strip(),
                str(row.get("Title", "")).strip(),
            )
            feedback_lookup[key] = (
                str(row.get("Action", "")).strip(),
                str(row.get("Notes", "")).strip(),
            )

        # Find File, Line, Title columns in the Excel for matching
        file_col = header_map.get("File")
        line_col = header_map.get("Line")
        title_col = header_map.get("Title")

        updated = 0
        for row_idx in range(2, ws.max_row + 1):
            file_val = str(ws.cell(row=row_idx, column=file_col).value or "").strip() if file_col else ""
            line_val = str(ws.cell(row=row_idx, column=line_col).value or "").strip() if line_col else ""
            title_val = str(ws.cell(row=row_idx, column=title_col).value or "").strip() if title_col else ""

            key = (file_val, line_val, title_val)
            if key in feedback_lookup:
                action, notes = feedback_lookup[key]
                ws.cell(row=row_idx, column=feedback_col, value=action)
                ws.cell(row=row_idx, column=constraints_col, value=notes)
                updated += 1

        wb.save(excel_path)
        logger.info("Updated %d rows in %s with feedback", updated, excel_path)
        return True

    except Exception as e:
        logger.error("Failed to save feedback to Excel: %s", e, exc_info=True)
        return False


def load_excel_as_dataframe(excel_path: str) -> Optional[pd.DataFrame]:
    """
    Load a detailed_code_review.xlsx (or customer-reported Excel) into a DataFrame
    suitable for the Review/Fix tabs.

    Handles column mapping:
        Feedback    → Action
        Constraints → Notes

    Returns None if the file cannot be loaded.
    """
    try:
        # Try to load the "Analysis" sheet first (primary issues sheet).
        # Fall back to first sheet if "Analysis" doesn't exist (e.g. customer Excel).
        try:
            df = pd.read_excel(excel_path, sheet_name="Analysis", engine="openpyxl")
        except ValueError:
            df = pd.read_excel(excel_path, sheet_name=0, engine="openpyxl")

        # Map Excel column names to UI column names
        rename_map = {}
        if "Feedback" in df.columns and "Action" not in df.columns:
            rename_map["Feedback"] = "Action"
        if "Constraints" in df.columns and "Notes" not in df.columns:
            rename_map["Constraints"] = "Notes"
        if rename_map:
            df = df.rename(columns=rename_map)

        # Drop S.No if present (UI doesn't use it)
        if "S.No" in df.columns:
            df = df.drop(columns=["S.No"])

        # Ensure Action and Notes columns exist
        if "Action" not in df.columns:
            df["Action"] = "Auto-fix"
        if "Notes" not in df.columns:
            df["Notes"] = ""

        # Fill NaN in Action with default
        df["Action"] = df["Action"].fillna("Auto-fix").replace("", "Auto-fix")

        return df

    except Exception as e:
        logger.error("Failed to load Excel %s: %s", excel_path, e)
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  Load deep analysis adapter sheets from Excel
# ═══════════════════════════════════════════════════════════════════════════════

def load_adapter_sheets(excel_path: str) -> Dict[str, pd.DataFrame]:
    """
    Load deep analysis adapter sheets (static_*) from detailed_code_review.xlsx.

    Returns a dict mapping sheet display names to DataFrames.
    Empty dict if no adapter sheets found or file cannot be loaded.
    """
    result: Dict[str, pd.DataFrame] = {}
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        for sheet_name in xls.sheet_names:
            if sheet_name.startswith("static_"):
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if not df.empty:
                    # Convert sheet name to display name: static_lizard → Lizard
                    display_name = sheet_name.replace("static_", "").replace("_", " ").title()
                    result[display_name] = df
    except Exception as e:
        logger.warning("Could not load adapter sheets from %s: %s", excel_path, e)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  DataFrame → Excel bytes (for download)
# ═══════════════════════════════════════════════════════════════════════════════

def export_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Analysis") -> bytes:
    """
    Export a DataFrame to styled Excel bytes suitable for st.download_button.

    Uses the CARE ExcelWriter if available, falls back to plain openpyxl.

    Returns:
        Bytes of the Excel workbook.
    """
    try:
        from utils.common.excel_writer import ExcelWriter, ExcelStyle

        buffer = io.BytesIO()
        temp_path = buffer  # CARE ExcelWriter needs a path, use temp file
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        writer = ExcelWriter(file_path=tmp_path)
        headers = list(df.columns)
        # Sanitize values: convert non-primitive types to str for openpyxl
        data_rows = [
            [str(v) if v is not None and not isinstance(v, (str, int, float, bool)) else v
             for v in row]
            for row in df.values.tolist()
        ]

        writer.add_table_sheet(
            headers=headers,
            data_rows=data_rows,
            sheet_name=sheet_name,
            status_column="Severity" if "Severity" in headers else None,
            autofit=True,
        )
        writer.save()

        with open(tmp_path, "rb") as f:
            excel_bytes = f.read()

        os.unlink(tmp_path)
        return excel_bytes

    except ImportError:
        # Fallback: plain pandas to_excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)
        return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  QA Traceability Report
# ═══════════════════════════════════════════════════════════════════════════════

def build_qa_traceability_report(
    feedback_df: pd.DataFrame,
    fixer_results: Optional[Dict[str, Any]] = None,
    audit_report_path: Optional[str] = None,
) -> pd.DataFrame:
    """
    Build a traceability report mapping human feedback decisions
    to final code modifications.

    Args:
        feedback_df: The reviewed DataFrame with Action/Notes columns.
        fixer_results: Results dict from CodebaseFixerAgent.run_agent().
        audit_report_path: Path to final_execution_audit.xlsx (optional).

    Returns:
        DataFrame with columns: File, Issue, Human_Action, Human_Notes,
        Fixer_Status, Lines_Changed, Validation_Status.
    """
    rows = []

    # Load audit report if available
    audit_data = {}
    if audit_report_path and os.path.exists(audit_report_path):
        try:
            audit_df = pd.read_excel(audit_report_path, sheet_name=0)
            for _, arow in audit_df.iterrows():
                key = (
                    str(arow.get("File", arow.get("file_path", ""))),
                    str(arow.get("Title", arow.get("issue_type", ""))),
                )
                audit_data[key] = {
                    "status": str(arow.get("Status", arow.get("status", "UNKNOWN"))),
                    "lines_changed": arow.get("Lines_Changed", arow.get("lines_changed", 0)),
                }
        except Exception as e:
            logger.warning("Could not load audit report: %s", e)

    # Also check fixer_results dict for per-file status
    fixer_file_status = {}
    if fixer_results and isinstance(fixer_results, dict):
        for item in fixer_results.get("results", []):
            if isinstance(item, dict):
                key = (str(item.get("file_path", "")), str(item.get("issue_type", "")))
                fixer_file_status[key] = item.get("status", "UNKNOWN")

    for _, row in feedback_df.iterrows():
        file_name = str(row.get("File", ""))
        issue_title = str(row.get("Title", row.get("Category", "")))
        action = str(row.get("Action", "Auto-fix"))
        notes = str(row.get("Notes", ""))

        # Look up fixer result
        lookup_key = (file_name, issue_title)
        audit_info = audit_data.get(lookup_key, {})
        fixer_status = (
            audit_info.get("status")
            or fixer_file_status.get(lookup_key, "PENDING")
        )

        rows.append({
            "File": file_name,
            "Issue": issue_title,
            "Severity": str(row.get("Severity", "")),
            "Human_Action": action,
            "Human_Notes": notes,
            "Fixer_Status": fixer_status,
            "Lines_Changed": audit_info.get("lines_changed", 0),
        })

    if not rows:
        return pd.DataFrame(columns=[
            "File", "Issue", "Severity", "Human_Action",
            "Human_Notes", "Fixer_Status", "Lines_Changed",
        ])

    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
#  Summary statistics
# ═══════════════════════════════════════════════════════════════════════════════

def compute_summary_stats(df: pd.DataFrame) -> Dict[str, Any]:
    """Compute summary statistics from an analysis results DataFrame."""
    total = len(df)

    # Normalize severity to Title Case for consistent matching
    severity_counts: Dict[str, int] = {}
    if "Severity" in df.columns:
        normalised = df["Severity"].astype(str).str.strip().str.title()
        severity_counts = normalised.value_counts().to_dict()

    action_counts: Dict[str, int] = {}
    if "Action" in df.columns:
        action_counts = df["Action"].value_counts().to_dict()

    to_fix = action_counts.get("Auto-fix", 0) + action_counts.get("Review", 0)
    to_skip = action_counts.get("Skip", 0)

    category_counts: Dict[str, int] = {}
    if "Category" in df.columns:
        category_counts = df["Category"].value_counts().to_dict()

    return {
        "total": total,
        "critical": severity_counts.get("Critical", 0),
        "high": severity_counts.get("High", 0),
        "medium": severity_counts.get("Medium", 0),
        "low": severity_counts.get("Low", 0),
        "to_fix": to_fix,
        "to_skip": to_skip,
        "to_review": action_counts.get("Review", 0) + action_counts.get("Manual Review", 0),
        "categories": category_counts,
        "unique_files": df["File"].nunique() if "File" in df.columns else 0,
    }
