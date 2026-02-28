"""
app.py

CARE — Codebase Analysis & Repair Engine
Streamlit dashboard for codebase ingestion, real-time analysis pipeline,
human-in-the-loop review, agentic code repair, and interactive chat.

Author: Pavan R
"""

import os
import re
import io
import json
import logging
import shutil
import time
import threading
from pathlib import Path
from queue import Queue, Empty
from typing import Dict, Optional

import streamlit as st
import pandas as pd

# --- Agent imports (graceful fallback) ---
try:
    from agents.codebase_analysis_chat_agent import (
        CodebaseAnalysisSessionState,
        CodebaseAnalysisOrchestration,
    )
    CHAT_AGENT_AVAILABLE = True
except ImportError:
    CHAT_AGENT_AVAILABLE = False

# --- Background workers ---
try:
    from ui.background_workers import (
        run_analysis_background,
        run_fixer_background,
        run_patch_analysis_background,
        ANALYSIS_PHASES,
        FIXER_PHASES,
        PATCH_PHASES,
    )
    WORKERS_AVAILABLE = True
except ImportError:
    WORKERS_AVAILABLE = False

# --- Feedback & QA helpers ---
try:
    from ui.feedback_helpers import (
        results_to_dataframe,
        dataframe_to_directives,
        export_to_excel_bytes,
        build_qa_traceability_report,
        compute_summary_stats,
        save_feedback_to_excel,
        load_excel_as_dataframe,
        load_adapter_sheets,
    )
    FEEDBACK_HELPERS_AVAILABLE = True
except ImportError:
    FEEDBACK_HELPERS_AVAILABLE = False

try:
    from ui.qa_inspector import QAInspector, create_zip_archive
    QA_INSPECTOR_AVAILABLE = True
except ImportError:
    QA_INSPECTOR_AVAILABLE = False

# --- Config imports ---
try:
    from utils.parsers.global_config_parser import GlobalConfig
    _gc = GlobalConfig()
    STREAMLIT_MODEL = _gc.get("llm.streamlit_model") or "qgenie::qwen2.5-14b-1m"
except Exception:
    try:
        from utils.parsers.env_parser import EnvConfig
        _ec = EnvConfig()
        STREAMLIT_MODEL = _ec.get("STREAMLIT_MODEL") or "qgenie::qwen2.5-14b-1m"
    except Exception:
        STREAMLIT_MODEL = "qgenie::qwen2.5-14b-1m"

import ui.streamlit_tools as st_tools

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────
APP_TITLE = "CARE — Codebase Analysis & Repair Engine"
APP_ICON = "🔬"
PLACEHOLDER = "🔬 _Analyzing..._"

# ── Logo paths ───────────────────────────────────────────────────────────────
_UI_DIR = os.path.dirname(__file__)
LOGO_MAIN = os.path.join(_UI_DIR, "qualcomm_logo.png")
LOGO_SIDEBAR = os.path.join(_UI_DIR, "qualcomm_logo_2.png")

# ── Page config (must be first Streamlit call) ───────────────────────────────
st.set_page_config(
    page_title=APP_TITLE,
    page_icon=APP_ICON,
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Session state defaults ───────────────────────────────────────────────────
_DEFAULTS = {
    # Chat (existing)
    "chat_history": [],
    "chat_summary": "",
    "all_feedback": [],
    "feedback_mode": False,
    "debug_mode": False,
    "active_page": "Analyze",
    # Ingestion & analysis
    "analysis_mode": "LLM Code Review",
    "codebase_path": "",
    "output_dir": "./out",
    "dependency_granularity": "File",
    "max_files": 2000,
    "batch_size": 25,
    "use_llm": True,
    "enable_adapters": False,
    "use_verible": False,
    "exclude_dirs": "",
    "exclude_globs": "",
    "exclude_headers": "",
    "custom_constraints": "",
    "file_to_fix": None,
    # Pipeline state
    "analysis_in_progress": False,
    "analysis_complete": False,
    "analysis_results": [],
    "analysis_metrics": {},
    "pipeline_logs": [],
    "phase_statuses": {},
    "analysis_thread": None,
    "log_queue": None,
    "result_store": {},
    # Review & feedback
    "feedback_df": None,
    # Patch analysis
    "patch_original_file": "",
    "patch_diff_file": "",
    "patch_modified_file": "",
    # Fixer state
    "fixer_in_progress": False,
    "fixer_complete": False,
    "fixer_logs": [],
    "fixer_phase_statuses": {},
    "fixer_thread": None,
    "fixer_log_queue": None,
    "fixer_result_store": {},
    "directives_path": None,
    # QA
    "qa_results": [],
    # Fixer audit feedback
    "_audit_feedback_saved": None,
    # Constraints generator
    "constraints_generated_md": "",
    # Configuration toggles
    "enable_hitl": False,
    "enable_telemetry": True,
}
for key, default in _DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = default


# ── Agent cache ──────────────────────────────────────────────────────────────
@st.cache_resource
def _get_orchestrator():
    if CHAT_AGENT_AVAILABLE:
        return CodebaseAnalysisOrchestration()
    return None


orchestrator = _get_orchestrator()


# ═══════════════════════════════════════════════════════════════════════════════
#  Helper functions
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_content(answer):
    """Pull plain-text content out of various LLM response shapes."""
    if isinstance(answer, dict) and "content" in answer:
        return answer["content"]
    if isinstance(answer, str):
        stripped = answer.strip()
        if stripped.startswith("{"):
            try:
                parsed = json.loads(stripped)
                if isinstance(parsed, dict) and "content" in parsed:
                    return parsed["content"]
            except (json.JSONDecodeError, ValueError):
                pass
        return answer
    return str(answer)


def _render_markdown_with_tables(md_text: str):
    """Render markdown, and additionally show any embedded tables as DataFrames."""
    st.markdown(md_text, unsafe_allow_html=True)
    table_pattern = r"(\|[^\n]+\|\n(?:\|[:\-]+\|)+\n(?:\|.*\|\n?)+)"
    for match in re.finditer(table_pattern, md_text):
        try:
            df = pd.read_csv(io.StringIO(match.group(0)), sep="|", engine="python")
            df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
            st.dataframe(df, width="stretch")
        except Exception:
            pass


def _validate_codebase_path(path: str) -> tuple:
    """Validate that a path exists and contains Verilog/SystemVerilog HDL files."""
    p = Path(path)
    if not p.exists():
        return False, "Path does not exist."
    if not p.is_dir():
        return False, "Path is not a directory."
    hdl_exts = {".v", ".sv", ".svh", ".vh"}
    found = any(p.rglob(f"*{ext}") for ext in hdl_exts)
    if not found:
        return False, "No Verilog/SystemVerilog files found in directory."
    return True, "Valid HDL codebase."


def _clean_output_dir(output_dir: str) -> None:
    """Remove previous run artifacts from the output directory.

    Deletes known subdirectories and report files so that a fresh analysis
    starts from a clean slate.  The output directory itself is preserved.
    """
    if not output_dir or not os.path.isdir(output_dir):
        return

    # Subdirectories to wipe entirely
    # NOTE: _uploads is NOT included here — it is cleared at upload time
    # (before new files are saved) so the analysis can still read them.
    _SUBDIRS = [
        "prompt_dumps",
        "parseddata",
        "shelved_backups",
    ]
    for sub in _SUBDIRS:
        p = os.path.join(output_dir, sub)
        if os.path.isdir(p):
            shutil.rmtree(p, ignore_errors=True)

    # Individual files to remove
    _FILES = [
        "detailed_code_review.xlsx",
        "detailed_code_review.csv",
        "llm_analysis_metrics.jsonl",
        "debug.log",
        "final_execution_audit.xlsx",
        "_generated.patch",
    ]
    for fname in _FILES:
        p = os.path.join(output_dir, fname)
        if os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass


def _drain_log_queue(queue_obj: Queue, target_list: list) -> bool:
    """Drain all pending messages from a Queue into a list. Returns True if __DONE__ found."""
    done = False
    try:
        while True:
            entry = queue_obj.get_nowait()
            if isinstance(entry, dict) and entry.get("message") == "__DONE__":
                done = True
            else:
                target_list.append(entry)
    except Empty:
        pass
    return done


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Analyze (Ingestion & Configuration)
# ═══════════════════════════════════════════════════════════════════════════════

def page_analyze():
    """Codebase ingestion controls and analysis configuration."""
    st.markdown(
        "<h2 style='text-align:center; margin-top:-10px;'>"
        "Codebase Analysis</h2>",
        unsafe_allow_html=True,
    )

    if st.session_state.get("analysis_in_progress"):
        st.warning("Analysis is currently running. Check the **Pipeline** page for progress.")
        return

    if st.session_state.get("analysis_complete"):
        st.success("Analysis complete! Review results on the **Review** page.")
        if st.button("Start New Analysis"):
            for key in [
                "analysis_complete", "analysis_results", "analysis_metrics",
                "pipeline_logs", "phase_statuses", "feedback_df",
                "fixer_complete", "fixer_logs", "qa_results",
                "result_store", "directives_path", "file_to_fix",
            ]:
                st.session_state[key] = _DEFAULTS.get(key, None)
            st.rerun()
        return

    # ── Input mode ────────────────────────────────────────────────────────
    st.markdown("### Input")
    input_mode = st.radio(
        "Input Source",
        ["Local Folder", "Upload Files", "Patch Analysis"],
        horizontal=True,
        label_visibility="collapsed",
    )

    is_patch_mode = input_mode == "Patch Analysis"

    if input_mode == "Local Folder":
        codebase_path = st_tools.folder_browser(
            label="Codebase Path",
            default_path=st.session_state.get("codebase_path", "./codebase"),
            key="analyze_codebase_browser",
            show_files=True,
            file_extensions=[".v", ".sv", ".svh", ".vh"],
            help_text="Path to a Verilog/SystemVerilog project directory, or a single source file.",
        )

        # ── Single-file detection ────────────────────────────────────
        # If the user selected a single HDL file instead of a directory,
        # treat the parent as the codebase and store the file as file_to_fix.
        hdl_exts = {".v", ".sv", ".svh", ".vh"}
        if (
            codebase_path
            and os.path.isfile(codebase_path)
            and Path(codebase_path).suffix.lower() in hdl_exts
        ):
            st.session_state["file_to_fix"] = codebase_path
            codebase_path = str(Path(codebase_path).parent)
            st.info(
                f"🎯 **Single-file mode**: analyzing only "
                f"`{Path(st.session_state['file_to_fix']).name}` "
                f"within `{codebase_path}`"
            )
        else:
            st.session_state["file_to_fix"] = None

        st.session_state["codebase_path"] = codebase_path

        if codebase_path:
            valid, msg = _validate_codebase_path(codebase_path)
            if valid:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")

    elif input_mode == "Upload Files":
        uploaded = st.file_uploader(
            "Upload HDL files",
            accept_multiple_files=True,
            type=["v", "sv", "svh", "vh"],
        )
        if uploaded:
            # Clear previous uploads so old files don't accumulate across runs
            upload_dir = os.path.join(st.session_state["output_dir"], "_uploads")
            if os.path.isdir(upload_dir):
                shutil.rmtree(upload_dir, ignore_errors=True)
            os.makedirs(upload_dir, exist_ok=True)
            for f in uploaded:
                with open(os.path.join(upload_dir, f.name), "wb") as out:
                    out.write(f.getbuffer())
            st.session_state["codebase_path"] = upload_dir
            st.success(f"✅ {len(uploaded)} files uploaded to staging area.")

    else:
        # ── Patch Analysis inputs ─────────────────────────────────────
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>🩹</span>"
            "<span style='font-size:16px; font-weight:600;'>"
            "Analyze a patch/diff for newly introduced issues"
            "</span></div>",
            unsafe_allow_html=True,
        )

        # Original file
        patch_original = st_tools.folder_browser(
            label="Original Source File",
            default_path=st.session_state.get("codebase_path", ""),
            key="patch_original_browser",
            show_files=True,
            file_extensions=[".v", ".sv", ".svh", ".vh"],
            help_text="Path to the original (unpatched) source file.",
        )
        st.session_state["patch_original_file"] = patch_original

        if patch_original and os.path.isfile(patch_original):
            st.success(f"✅ Original: `{os.path.basename(patch_original)}`")

        # Patch / diff file
        patch_file_path = st_tools.folder_browser(
            label="Patch / Diff File",
            default_path=st.session_state.get("output_dir", ""),
            key="patch_diff_browser",
            show_files=True,
            file_extensions=[".patch", ".diff", ".txt"],
            help_text="Path to the .patch or .diff file (unified diff format).",
        )
        st.session_state["patch_diff_file"] = patch_file_path

        if patch_file_path and os.path.isfile(patch_file_path):
            st.success(f"✅ Patch: `{os.path.basename(patch_file_path)}`")

        # Modified file (optional) — if provided WITHOUT a patch file, diff is auto-generated
        st.caption(
            "**Optional**: Provide the modified file instead of (or in addition to) a patch file. "
            "If only the modified file is given, a diff will be generated automatically."
        )
        patch_modified = st_tools.folder_browser(
            label="Modified File (optional)",
            default_path="",
            key="patch_modified_browser",
            show_files=True,
            file_extensions=[".v", ".sv", ".svh", ".vh"],
            help_text="Path to the modified source file. If no patch file is provided, a diff will be generated.",
        )
        st.session_state["patch_modified_file"] = patch_modified

        if patch_modified and os.path.isfile(patch_modified):
            st.info(f"Modified: `{os.path.basename(patch_modified)}`")

        # Upload fallback for patch file
        if not (patch_file_path and os.path.isfile(patch_file_path)):
            uploaded_patch = st.file_uploader(
                "Or upload a patch/diff file",
                type=["patch", "diff", "txt"],
                key="patch_upload",
            )
            if uploaded_patch:
                out_dir = st.session_state.get("output_dir", "./out")
                os.makedirs(out_dir, exist_ok=True)
                saved_path = os.path.join(out_dir, uploaded_patch.name)
                with open(saved_path, "wb") as f:
                    f.write(uploaded_patch.getbuffer())
                st.session_state["patch_diff_file"] = saved_path
                patch_file_path = saved_path
                st.success(f"✅ Patch uploaded: `{uploaded_patch.name}`")

    st.divider()

    # ── Analysis configuration ────────────────────────────────────────────
    if not is_patch_mode:
        st.markdown("### Configuration")

        col1, col2 = st.columns(2)
        with col1:
            analysis_mode = st.selectbox(
                "Analysis Mode",
                ["LLM Code Review", "Static Analysis Only"],
                help=(
                    "**LLM Code Review**: Per-file semantic analysis using an LLM (produces Excel report).\n\n"
                    "**Static Analysis Only**: Fast regex-based 7-phase pipeline (produces health report JSON)."
                ),
            )
            st.session_state["analysis_mode"] = analysis_mode

        with col2:
            max_files = st.number_input("Max Files", min_value=1, max_value=50000, value=2000)
            st.session_state["max_files"] = max_files

            batch_size = st.number_input("Batch Size", min_value=1, max_value=200, value=25)
            st.session_state["batch_size"] = batch_size

    # Advanced options (shared by all modes)
    with st.expander("Advanced Options"):
        col1, col2 = st.columns(2)
        with col1:
            enable_adapters = st.checkbox(
                "Enable Deep Adapters (Lizard, Flawfinder)",
                value=False,
            )
            st.session_state["enable_adapters"] = enable_adapters

            if not is_patch_mode:
                use_verible = st.checkbox(
                    "Use Verible (semantic dependency analysis)",
                    value=False,
                    help="Requires verible installed. Provides accurate module hierarchy and dependency context for LLM analysis.",
                )
                st.session_state["use_verible"] = use_verible

                if use_verible:
                    dep_granularity = st.selectbox(
                        "Dependency Granularity",
                        ["File", "Module", "Package"],
                        help=(
                            "**File**: Individual source/header files.\n\n"
                            "**Module**: Group by directory (component-level).\n\n"
                            "**Package**: Top-level architecture layers."
                        ),
                    )
                    st.session_state["dependency_granularity"] = dep_granularity

        with col2:
            if not is_patch_mode:
                exclude_dirs = st.text_input(
                    "Exclude Directories (comma-separated)",
                    value="",
                    help="e.g., test,third_party,build",
                )
                st.session_state["exclude_dirs"] = exclude_dirs
                exclude_globs = st.text_input(
                    "Exclude Glob Patterns (comma-separated)",
                    value="",
                    help="e.g., */test/*, *_tb.sv, *_autogen/*",
                )
                st.session_state["exclude_globs"] = exclude_globs
                exclude_headers = st.text_input(
                    "Exclude Headers (comma-separated)",
                    value="",
                    help="Header files to exclude from context injection. Supports exact names, basenames, or glob patterns. e.g., auto_generated.svh, debug_*.vh",
                )
                st.session_state["exclude_headers"] = exclude_headers
                custom_constraints = st.text_input(
                    "Custom Constraint Files (comma-separated .md paths)",
                    value="",
                    help="e.g., my_rules.md, /abs/path/extra_constraints.md",
                )
                st.session_state["custom_constraints"] = custom_constraints

        output_dir = st_tools.folder_browser(
            label="Output Directory",
            default_path=st.session_state.get("output_dir", "./out"),
            key="analyze_output_browser",
            show_files=False,
            help_text="Directory where analysis reports will be saved.",
        )
        st.session_state["output_dir"] = output_dir

    st.divider()

    # ── Start analysis ────────────────────────────────────────────────────
    if is_patch_mode:
        # Patch mode validation
        has_original = bool(
            st.session_state.get("patch_original_file")
            and os.path.isfile(str(st.session_state.get("patch_original_file", "")))
        )
        has_patch = bool(
            st.session_state.get("patch_diff_file")
            and os.path.isfile(str(st.session_state.get("patch_diff_file", "")))
        )
        has_modified = bool(
            st.session_state.get("patch_modified_file")
            and os.path.isfile(str(st.session_state.get("patch_modified_file", "")))
        )
        can_start = has_original and (has_patch or has_modified)

        if not has_original:
            st.warning("Select the original source file.")
        elif not has_patch and not has_modified:
            st.warning("Provide either a patch/diff file or a modified file.")

        if st.button("🩹 Analyze Patch", type="primary", disabled=not can_start):
            if not WORKERS_AVAILABLE:
                st.error("Background workers module not available.")
                return

            config = {
                "file_path": st.session_state["patch_original_file"],
                "patch_file": st.session_state.get("patch_diff_file", ""),
                "modified_file": st.session_state.get("patch_modified_file", ""),
                "output_dir": st.session_state["output_dir"],
                "enable_adapters": st.session_state.get("enable_adapters", False),
                "debug_mode": st.session_state.get("debug_mode", False),
                "exclude_headers": [
                    h.strip()
                    for h in st.session_state.get("exclude_headers", "").split(",")
                    if h.strip()
                ],
            }

            # Clean previous run artifacts
            _clean_output_dir(config["output_dir"])

            log_queue = Queue()
            result_store = {"status": "running", "phase_statuses": {}}

            st.session_state["log_queue"] = log_queue
            st.session_state["result_store"] = result_store
            st.session_state["pipeline_logs"] = []
            st.session_state["phase_statuses"] = {i: "pending" for i in range(1, 8)}
            st.session_state["analysis_mode"] = "Patch Analysis"

            t = threading.Thread(
                target=run_patch_analysis_background,
                args=(config, log_queue, result_store),
                daemon=True,
            )
            t.start()
            st.session_state["analysis_thread"] = t
            st.session_state["analysis_in_progress"] = True
            st.rerun()
    else:
        # Standard analysis validation
        can_start = bool(st.session_state.get("codebase_path"))
        if can_start:
            valid, _ = _validate_codebase_path(st.session_state["codebase_path"])
            can_start = valid

        if st.button("🚀 Start Analysis", type="primary", disabled=not can_start):
            if not WORKERS_AVAILABLE:
                st.error("Background workers module not available. Check ui/background_workers.py.")
                return

            # Prepare config
            exclude = [
                d.strip()
                for d in st.session_state.get("exclude_dirs", "").split(",")
                if d.strip()
            ]
            exclude_globs = [
                g.strip()
                for g in st.session_state.get("exclude_globs", "").split(",")
                if g.strip()
            ]
            exclude_headers = [
                h.strip()
                for h in st.session_state.get("exclude_headers", "").split(",")
                if h.strip()
            ]
            config = {
                "codebase_path": st.session_state["codebase_path"],
                "output_dir": st.session_state["output_dir"],
                "analysis_mode": (
                    "llm_exclusive"
                    if st.session_state["analysis_mode"] == "LLM Code Review"
                    else "static"
                ),
                "dependency_granularity": st.session_state["dependency_granularity"],
                "use_llm": st.session_state["analysis_mode"] == "LLM Code Review",
                "enable_adapters": st.session_state.get("enable_adapters", False),
                "use_verible": st.session_state.get("use_verible", False),
                "enable_hitl": st.session_state.get("enable_hitl", False),
                "max_files": st.session_state.get("max_files", 2000),
                "batch_size": st.session_state.get("batch_size", 25),
                "exclude_dirs": exclude,
                "exclude_globs": exclude_globs,
                "exclude_headers": exclude_headers,
                "custom_constraints": [
                    c.strip()
                    for c in st.session_state.get("custom_constraints", "").split(",")
                    if c.strip()
                ],
                "debug_mode": st.session_state.get("debug_mode", False),
                "file_to_fix": st.session_state.get("file_to_fix"),
            }

            # Clean previous run artifacts from the output directory
            _clean_output_dir(config["output_dir"])

            # Initialize queue and result store
            log_queue = Queue()
            result_store = {"status": "running", "phase_statuses": {}}

            st.session_state["log_queue"] = log_queue
            st.session_state["result_store"] = result_store
            st.session_state["pipeline_logs"] = []
            st.session_state["phase_statuses"] = {i: "pending" for i in range(1, 8)}

            # Launch background thread
            t = threading.Thread(
                target=run_analysis_background,
                args=(config, log_queue, result_store),
                daemon=True,
            )
            t.start()
            st.session_state["analysis_thread"] = t
            st.session_state["analysis_in_progress"] = True
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Pipeline (Real-Time Progress Terminal)
# ═══════════════════════════════════════════════════════════════════════════════

def page_pipeline():
    """Real-time pipeline terminal showing analysis progress."""
    st.markdown(
        "<h2 style='text-align:center; margin-top:-10px;'>"
        "Analysis Pipeline</h2>",
        unsafe_allow_html=True,
    )

    in_progress = st.session_state.get("analysis_in_progress", False)
    is_complete = st.session_state.get("analysis_complete", False)

    if not in_progress and not is_complete:
        st.info("No analysis running. Go to **Analyze** to start one.")
        return

    # Phase tracker — pick the right set of phase labels
    is_patch = st.session_state.get("analysis_mode") == "Patch Analysis"
    if WORKERS_AVAILABLE:
        phases = PATCH_PHASES if is_patch else ANALYSIS_PHASES
    else:
        phases = {i: f"Phase {i}" for i in range(1, 8)}
    phase_statuses = st.session_state.get("phase_statuses", {})

    # Update from result_store (thread-safe read)
    result_store = st.session_state.get("result_store", {})
    if "phase_statuses" in result_store:
        phase_statuses.update(result_store["phase_statuses"])
        st.session_state["phase_statuses"] = phase_statuses

    st_tools.render_phase_tracker(phases, phase_statuses)
    st.divider()

    # Drain log queue
    log_queue = st.session_state.get("log_queue")
    if log_queue:
        done = _drain_log_queue(log_queue, st.session_state["pipeline_logs"])
        if done and in_progress:
            st.session_state["analysis_in_progress"] = False
            st.session_state["analysis_complete"] = True

            # Harvest results from shared store
            if result_store.get("status") == "success":
                st.session_state["analysis_results"] = result_store.get("analysis_results", [])
                st.session_state["analysis_metrics"] = result_store.get("analysis_metrics", {})
                # Invalidate cached feedback DataFrame so Review tab rebuilds it
                st.session_state["feedback_df"] = None
            st.rerun()

    # Log stream
    st.markdown("### Console Output")
    st_tools.render_log_stream(st.session_state.get("pipeline_logs", []))

    if in_progress:
        # Auto-refresh while running
        time.sleep(0.5)
        st.rerun()

    if is_complete:
        st.divider()
        st.success("✅ Analysis complete!")

        # Summary metrics
        results = st.session_state.get("analysis_results", [])
        metrics = st.session_state.get("analysis_metrics", {})

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Issues Found", len(results))
        with col2:
            overall = metrics.get("overall_health", {})
            score = overall.get("score", "N/A") if isinstance(overall, dict) else "N/A"
            st.metric("Health Score", f"{score}/100" if score != "N/A" else "N/A")
        with col3:
            stats = metrics.get("statistics", {})
            files = stats.get("processed_files", stats.get("total_files", len(results)))
            st.metric("Files Analyzed", files)

        col1, col2 = st.columns(2)
        with col1:
            report_path = result_store.get("report_path") or result_store.get("health_report_path")
            if report_path and os.path.exists(str(report_path)):
                with open(str(report_path), "rb") as f:
                    st.download_button(
                        "📥 Download Report",
                        data=f.read(),
                        file_name=os.path.basename(str(report_path)),
                    )
        with col2:
            if results:
                st.success("✅ Analysis complete — switch to the **Review** tab to inspect results.")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Review & Feedback (HITL)
# ═══════════════════════════════════════════════════════════════════════════════

def page_review():
    """Interactive review of analysis results with editable feedback columns."""
    st.markdown(
        "<div style='display:flex; align-items:center; gap:12px; margin-bottom:16px;'>"
        "<span style='font-size:36px;'>📋</span>"
        "<div>"
        "<h2 style='margin:0; padding:0;'>Review & Feedback</h2>"
        "<span style='color:#888; font-size:14px;'>Inspect findings, set actions, and export directives</span>"
        "</div></div>",
        unsafe_allow_html=True,
    )

    if not FEEDBACK_HELPERS_AVAILABLE:
        st.error("Feedback helpers module not available.")
        return

    results = st.session_state.get("analysis_results", [])
    out_dir = st.session_state.get("output_dir", "./out")
    review_excel = os.path.join(out_dir, "detailed_code_review.xlsx")

    # Build DataFrame if not cached — prefer Excel (authoritative), fall back to in-memory results
    if st.session_state.get("feedback_df") is None:
        if os.path.isfile(review_excel):
            st.session_state["feedback_df"] = load_excel_as_dataframe(review_excel)
        if st.session_state.get("feedback_df") is None and results:
            st.session_state["feedback_df"] = results_to_dataframe(results)

    # Reload button
    reload_col, info_col = st.columns([1, 4])
    with reload_col:
        if st.button("🔄 Reload from Excel", key="review_reload"):
            if os.path.isfile(review_excel):
                st.session_state["feedback_df"] = load_excel_as_dataframe(review_excel)
                st.rerun()
            else:
                st.warning("No `detailed_code_review.xlsx` found.")
    with info_col:
        if st.session_state.get("feedback_df") is not None:
            src = "Excel" if os.path.isfile(review_excel) else "in-memory results"
            st.caption(f"Loaded from {src} — {len(st.session_state['feedback_df'])} issues")

    df = st.session_state.get("feedback_df")
    if df is None or df.empty:
        st.info(
            "No analysis results available. Run an analysis from the **Analyze** tab first."
        )
        return

    # Ensure editable text columns are string-typed (NaN floats break data_editor)
    for _col in ("Notes", "Action"):
        if _col in df.columns:
            df[_col] = df[_col].fillna("").astype(str)

    # Normalize severity values to title case for consistent counting
    if "Severity" in df.columns:
        df["Severity"] = df["Severity"].astype(str).str.strip().str.title()

    # ── Filters ───────────────────────────────────────────────────────────
    with st.expander("🔍 Filters", expanded=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            sev_options = sorted(df["Severity"].unique().tolist()) if "Severity" in df.columns else []
            sev_filter = st.multiselect("Severity", sev_options, default=sev_options)
        with col2:
            cat_options = sorted(df["Category"].unique().tolist()) if "Category" in df.columns else []
            cat_filter = st.multiselect("Category", cat_options, default=cat_options)
        with col3:
            file_options = sorted(df["File"].unique().tolist()) if "File" in df.columns else []
            file_filter = st.multiselect("File", file_options, default=file_options)

    # Apply filters
    mask = pd.Series([True] * len(df), index=df.index)
    if sev_filter and "Severity" in df.columns:
        mask &= df["Severity"].isin(sev_filter)
    if cat_filter and "Category" in df.columns:
        mask &= df["Category"].isin(cat_filter)
    if file_filter and "File" in df.columns:
        mask &= df["File"].isin(file_filter)

    filtered = df[mask].copy()

    # ── Summary metrics (modern cards) ────────────────────────────────────
    stats = compute_summary_stats(filtered)

    def _metric_card(label, value, color, icon):
        return (
            f"<div style='background:{st_tools.CARE_SURFACE}; border:1px solid {st_tools.CARE_BORDER}; "
            f"border-radius:12px; padding:16px 20px; text-align:center;'>"
            f"<div style='font-size:22px; margin-bottom:4px;'>{icon}</div>"
            f"<div style='font-size:28px; font-weight:700; color:{color};'>{value}</div>"
            f"<div style='font-size:12px; color:{st_tools.CARE_TEXT_SECONDARY}; font-weight:500; "
            f"text-transform:uppercase; letter-spacing:0.5px; margin-top:2px;'>{label}</div>"
            f"</div>"
        )

    mcols = st.columns(7)
    cards = [
        ("Total Issues", stats["total"], st_tools.CARE_TEXT, "📊"),
        ("Critical", stats["critical"], st_tools.CARE_RED, "🔴"),
        ("High", stats["high"], "#FF6B35", "🟠"),
        ("Medium", stats["medium"], st_tools.CARE_GOLD, "🟡"),
        ("Low", stats["low"], st_tools.CARE_GREEN, "🟢"),
        ("To Fix", stats["to_fix"], st_tools.CARE_PRIMARY, "🔧"),
        ("To Skip", stats["to_skip"], st_tools.CARE_TEXT_SECONDARY, "⏭️"),
    ]
    for i, (label, value, color, icon) in enumerate(cards):
        with mcols[i]:
            st.markdown(_metric_card(label, value, color, icon), unsafe_allow_html=True)

    # File coverage
    unique_files = stats.get("unique_files", 0)
    st.markdown(
        f"<div style='text-align:center; margin:12px 0 4px 0; color:{st_tools.CARE_TEXT_SECONDARY}; font-size:13px;'>"
        f"📁 <b>{unique_files}</b> unique file{'s' if unique_files != 1 else ''} affected"
        f"</div>",
        unsafe_allow_html=True,
    )

    st.divider()

    # ── Editable data table ───────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📝</span>"
        "<span style='font-size:18px; font-weight:600;'>Issue Table</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "Edit the **Action** and **Notes** columns to provide feedback. "
        "Choose **Skip** to ignore false positives, **Auto-fix** to apply the suggested fix, "
        "or **Review** to flag for manual inspection with custom notes."
    )

    column_config = {
        "Action": st.column_config.SelectboxColumn(
            "Action",
            options=["Auto-fix", "Skip", "Review"],
            default="Auto-fix",
            width="small",
        ),
        "Notes": st.column_config.TextColumn(
            "Notes",
            help="Add constraints or feedback for the fixer agent.",
            width="medium",
        ),
        "Severity": st.column_config.TextColumn("Severity", width="small"),
        "Confidence": st.column_config.TextColumn("Confidence", width="small"),
        "Line": st.column_config.NumberColumn("Line", width="small"),
    }

    edited = st.data_editor(
        filtered,
        width="stretch",
        column_config=column_config,
        disabled=[
            "File", "Title", "Severity", "Confidence", "Category",
            "Line", "Description", "Suggestion", "Code", "Fixed_Code",
        ],
        hide_index=True,
        key="review_editor",
    )

    # Write edits back to the full DataFrame
    if edited is not None:
        for col in ["Action", "Notes"]:
            if col in edited.columns:
                df.loc[edited.index, col] = edited[col]
        st.session_state["feedback_df"] = df

    st.divider()

    # ── Save feedback to Excel ──────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>💾</span>"
        "<span style='font-size:18px; font-weight:600;'>Save Feedback</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "Save your **Action** and **Notes** feedback back to the "
        "`detailed_code_review.xlsx` file in the output directory."
    )

    out_dir = st.session_state.get("output_dir", "./out")
    review_excel_path = os.path.join(out_dir, "detailed_code_review.xlsx")

    save_col1, save_col2 = st.columns([1, 3])
    with save_col1:
        if st.button("💾 Save Feedback to Excel", type="primary"):
            ok = save_feedback_to_excel(df, review_excel_path)
            if ok:
                st.session_state["_feedback_saved"] = True
            else:
                st.session_state["_feedback_saved"] = False

    with save_col2:
        if st.session_state.get("_feedback_saved") is True:
            st.success(f"Feedback saved to `{review_excel_path}`")
        elif st.session_state.get("_feedback_saved") is False:
            st.error("Failed to save feedback. Check logs for details.")

    st.divider()

    # ── Deep Analysis Reports (adapter sheets) ────────────────────────────
    if os.path.isfile(review_excel_path):
        adapter_sheets = load_adapter_sheets(review_excel_path)
        if adapter_sheets:
            st.markdown(
                "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
                "<span style='font-size:20px;'>🔬</span>"
                "<span style='font-size:18px; font-weight:600;'>Deep Analysis Reports</span>"
                "</div>",
                unsafe_allow_html=True,
            )
            st.caption("Results from deep static analysis adapters (Lizard, Flawfinder, CCLS).")

            for adapter_name, adapter_df in adapter_sheets.items():
                with st.expander(f"**{adapter_name}** — {len(adapter_df)} findings", expanded=False):
                    # Summary metrics row
                    if "Severity" in adapter_df.columns:
                        sev_counts = adapter_df["Severity"].value_counts()
                        metric_cols = st.columns(min(len(sev_counts), 5))
                        for i, (sev, count) in enumerate(sev_counts.items()):
                            if i < len(metric_cols):
                                metric_cols[i].metric(str(sev).title(), count)

                    st.dataframe(
                        adapter_df,
                        width="stretch",
                        hide_index=True,
                        height=min(400, 35 * (len(adapter_df) + 1)),
                    )

            st.divider()

    # ── Downloads & proceed ───────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📦</span>"
        "<span style='font-size:18px; font-weight:600;'>Export & Proceed</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    col1, col2, col3 = st.columns(3)

    with col1:
        excel_bytes = export_to_excel_bytes(df)
        st.download_button(
            "📥 Download Excel",
            data=excel_bytes,
            file_name="cure_analysis_review.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col2:
        # Sanitize non-serializable values before JSON export
        export_df = df.copy()
        for col_name in export_df.columns:
            export_df[col_name] = export_df[col_name].apply(
                lambda v: str(v) if v is not None and not isinstance(v, (str, int, float, bool)) else v
            )
        json_str = export_df.to_json(orient="records", indent=2)
        st.download_button(
            "📥 Download JSON",
            data=json_str,
            file_name="cure_analysis_review.json",
            mime="application/json",
        )
    with col3:
        if st.button("⚡ Proceed to Fix & QA", type="primary"):
            directives_path = os.path.join(out_dir, "agent_directives.jsonl")
            dataframe_to_directives(df, directives_path)
            st.session_state["directives_path"] = directives_path
            st.success("✅ Directives saved — switch to the **Fix & QA** tab.")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Fix & QA
# ═══════════════════════════════════════════════════════════════════════════════

def page_fixer_qa():
    """Apply fixes based on feedback and run QA validation."""
    st.markdown(
        "<div style='display:flex; align-items:center; gap:12px; margin-bottom:16px;'>"
        "<span style='font-size:36px;'>🔧</span>"
        "<div>"
        "<h2 style='margin:0; padding:0;'>Fix & QA Validation</h2>"
        "<span style='color:#888; font-size:14px;'>Load issues, apply fixes, and validate results</span>"
        "</div></div>",
        unsafe_allow_html=True,
    )

    fixer_in_progress = st.session_state.get("fixer_in_progress", False)
    fixer_complete = st.session_state.get("fixer_complete", False)
    out_dir = st.session_state.get("output_dir", "./out")

    # ── Section 1: Issue Source Selection ─────────────────────────────────
    if not fixer_in_progress and not fixer_complete:
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>📂</span>"
            "<span style='font-size:18px; font-weight:600;'>1. Load Issues</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        # Auto-detect detailed_code_review.xlsx in the output folder
        auto_excel_path = os.path.join(out_dir, "detailed_code_review.xlsx")
        has_auto_excel = os.path.isfile(auto_excel_path)
        has_directives = (
            st.session_state.get("directives_path")
            and os.path.exists(str(st.session_state.get("directives_path", "")))
        )

        source_options = []
        if has_directives:
            source_options.append("From Review tab (directives ready)")
        if has_auto_excel:
            source_options.append(f"Auto-detected: {os.path.basename(auto_excel_path)}")
        source_options.append("Browse for Excel file (customer-reported issues)")
        source_options.append("Upload directives JSONL file")

        source_choice = st.radio(
            "Select issue source",
            source_options,
            index=0,
            key="fixer_source_choice",
        )

        directives_path = None
        fixer_df = None

        # --- From Review tab ---
        if source_choice.startswith("From Review tab"):
            directives_path = st.session_state.get("directives_path")
            st.success(f"Using directives: `{directives_path}`")

        # --- Auto-detected Excel ---
        elif source_choice.startswith("Auto-detected"):
            if FEEDBACK_HELPERS_AVAILABLE:
                fixer_df = load_excel_as_dataframe(auto_excel_path)
                if fixer_df is not None:
                    st.success(f"Loaded `{auto_excel_path}` — **{len(fixer_df)}** issues found")
                    with st.expander("Preview loaded issues", expanded=False):
                        st.dataframe(fixer_df.head(20), width="stretch", hide_index=True)
                else:
                    st.error(f"Failed to load `{auto_excel_path}`")
                    return

        # --- Browse for customer Excel ---
        elif source_choice.startswith("Browse for Excel"):
            browse_path = st_tools.folder_browser(
                label="Select Excel file with issues",
                default_path=out_dir,
                key="fixer_excel_browser",
                show_files=True,
                help_text="Browse for a customer-reported .xlsx file with issues to fix.",
            )
            # Check if the selected path is an Excel file
            if browse_path and os.path.isfile(browse_path) and browse_path.endswith((".xlsx", ".xls")):
                if FEEDBACK_HELPERS_AVAILABLE:
                    fixer_df = load_excel_as_dataframe(browse_path)
                    if fixer_df is not None:
                        st.success(f"Loaded `{browse_path}` — **{len(fixer_df)}** issues found")
                        with st.expander("Preview loaded issues", expanded=False):
                            st.dataframe(fixer_df.head(20), width="stretch", hide_index=True)
                    else:
                        st.error(f"Failed to load `{browse_path}`")
                        return
            elif browse_path:
                # Also allow uploading via file_uploader as fallback
                uploaded_excel = st.file_uploader(
                    "Or upload an Excel file",
                    type=["xlsx", "xls"],
                    key="fixer_excel_upload",
                )
                if uploaded_excel:
                    upload_path = os.path.join(out_dir, uploaded_excel.name)
                    os.makedirs(out_dir, exist_ok=True)
                    with open(upload_path, "wb") as f:
                        f.write(uploaded_excel.getbuffer())
                    if FEEDBACK_HELPERS_AVAILABLE:
                        fixer_df = load_excel_as_dataframe(upload_path)
                        if fixer_df is not None:
                            st.success(f"Uploaded and loaded — **{len(fixer_df)}** issues")
                        else:
                            st.error("Failed to parse the uploaded Excel.")
                            return
                else:
                    st.info("Select an `.xlsx` file from the browser above, or upload one.")
                    return

        # --- Upload JSONL ---
        elif source_choice.startswith("Upload directives"):
            uploaded = st.file_uploader(
                "Upload a directives JSONL file",
                type=["jsonl", "json"],
                key="fixer_jsonl_upload",
            )
            if uploaded:
                os.makedirs(out_dir, exist_ok=True)
                directives_path = os.path.join(out_dir, "agent_directives.jsonl")
                with open(directives_path, "wb") as f:
                    f.write(uploaded.getbuffer())
                st.session_state["directives_path"] = directives_path
                st.success(f"Directives uploaded: `{directives_path}`")
            else:
                return

        # If we loaded a DataFrame (Excel source), convert to directives
        if fixer_df is not None and directives_path is None:
            if FEEDBACK_HELPERS_AVAILABLE:
                os.makedirs(out_dir, exist_ok=True)
                directives_path = os.path.join(out_dir, "agent_directives.jsonl")
                dataframe_to_directives(fixer_df, directives_path)
                st.session_state["directives_path"] = directives_path
                # Also store the DataFrame for traceability
                if st.session_state.get("feedback_df") is None:
                    st.session_state["feedback_df"] = fixer_df

        if not directives_path:
            return

        st.divider()

        # ── Fixer configuration ──────────────────────────────────────────
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>⚙️</span>"
            "<span style='font-size:18px; font-weight:600;'>2. Fixer Configuration</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        output_dir = st_tools.folder_browser(
            label="Output Directory for Fixed Files",
            default_path=out_dir,
            key="fixer_output_browser",
            show_files=False,
            help_text="Directory where fixed source files will be written.",
        )
        dry_run = st.checkbox("Dry Run (simulate without writing)", value=False)

        st.divider()

        if st.button("🔧 Apply Fixes", type="primary"):
            if not WORKERS_AVAILABLE:
                st.error("Background workers not available.")
                return

            config = {
                "directives_path": directives_path,
                "codebase_path": st.session_state.get("codebase_path", "./codebase"),
                "output_dir": output_dir,
                "dry_run": dry_run,
                "debug_mode": st.session_state.get("debug_mode", False),
            }

            fixer_queue = Queue()
            fixer_result_store = {"fixer_status": "running"}

            st.session_state["fixer_log_queue"] = fixer_queue
            st.session_state["fixer_result_store"] = fixer_result_store
            st.session_state["fixer_logs"] = []
            st.session_state["fixer_phase_statuses"] = {i: "pending" for i in range(1, 5)}

            t = threading.Thread(
                target=run_fixer_background,
                args=(config, fixer_queue, fixer_result_store),
                daemon=True,
            )
            t.start()
            st.session_state["fixer_thread"] = t
            st.session_state["fixer_in_progress"] = True
            st.rerun()

    # ── Section 3: Execution log ─────────────────────────────────────────
    if fixer_in_progress:
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>📋</span>"
            "<span style='font-size:18px; font-weight:600;'>3. Execution Log</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        phases = FIXER_PHASES if WORKERS_AVAILABLE else {i: f"Phase {i}" for i in range(1, 5)}
        fixer_result_store = st.session_state.get("fixer_result_store", {})
        fixer_statuses = st.session_state.get("fixer_phase_statuses", {})

        if "fixer_phase_statuses" in fixer_result_store:
            fixer_statuses.update(fixer_result_store["fixer_phase_statuses"])
            st.session_state["fixer_phase_statuses"] = fixer_statuses

        st_tools.render_phase_tracker(phases, fixer_statuses)
        st.divider()

        # Drain fixer log queue
        fixer_queue = st.session_state.get("fixer_log_queue")
        if fixer_queue:
            done = _drain_log_queue(fixer_queue, st.session_state["fixer_logs"])
            if done:
                st.session_state["fixer_in_progress"] = False
                st.session_state["fixer_complete"] = True
                st.rerun()

        st_tools.render_log_stream(st.session_state.get("fixer_logs", []))
        time.sleep(0.5)
        st.rerun()

    # ── Section 4: QA Validation ─────────────────────────────────────────
    if fixer_complete:
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>✅</span>"
            "<span style='font-size:18px; font-weight:600;'>4. QA Validation Results</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        # Run QA if not already done
        if not st.session_state.get("qa_results") and QA_INSPECTOR_AVAILABLE:
            with st.spinner("Running QA validation..."):
                output_dir = st.session_state.get("output_dir", "./out")
                codebase_path = st.session_state.get("codebase_path", "./codebase")

                inspector = QAInspector(
                    fixed_codebase_path=codebase_path,
                    original_results=st.session_state.get("analysis_results", []),
                    original_metrics=st.session_state.get("analysis_metrics", {}),
                )
                qa_results = inspector.validate_all()
                st.session_state["qa_results"] = qa_results

        qa_results = st.session_state.get("qa_results", [])
        if qa_results:
            qa_df = pd.DataFrame(qa_results)
            st.dataframe(
                qa_df.style.apply(
                    lambda row: [
                        "background-color: #F0FDF4" if row.get("Pass") else "background-color: #FEF2F2"
                    ] * len(row),
                    axis=1,
                ),
                width="stretch",
                hide_index=True,
            )

            # Summary
            passed = sum(1 for r in qa_results if r.get("Pass"))
            failed = sum(1 for r in qa_results if not r.get("Pass"))
            st.markdown(
                f"**QA Summary:** {passed} checks passed, {failed} checks failed."
            )
        else:
            st.info("QA Inspector not available or no results.")

        st.divider()

        # ── Section 5: Traceability report ───────────────────────────────
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>📊</span>"
            "<span style='font-size:18px; font-weight:600;'>5. Traceability Report</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        feedback_df = st.session_state.get("feedback_df")
        fixer_result_store = st.session_state.get("fixer_result_store", {})
        fixer_results = fixer_result_store.get("fixer_results")
        audit_path = fixer_result_store.get("audit_report_path")

        if feedback_df is not None and FEEDBACK_HELPERS_AVAILABLE:
            trace_df = build_qa_traceability_report(
                feedback_df,
                fixer_results=fixer_results,
                audit_report_path=audit_path,
            )
            st.dataframe(trace_df, width="stretch", hide_index=True)

            # Downloads
            col1, col2, col3 = st.columns(3)
            with col1:
                trace_excel = export_to_excel_bytes(trace_df, sheet_name="QA Traceability")
                st.download_button(
                    "📥 Download QA Report (Excel)",
                    data=trace_excel,
                    file_name="cure_qa_traceability.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with col2:
                trace_json = trace_df.to_json(orient="records", indent=2)
                st.download_button(
                    "📥 Download QA Report (JSON)",
                    data=trace_json,
                    file_name="cure_qa_traceability.json",
                    mime="application/json",
                )
            with col3:
                if QA_INSPECTOR_AVAILABLE:
                    codebase_path = st.session_state.get("codebase_path", "")
                    if codebase_path and os.path.isdir(codebase_path):
                        output_dir = st.session_state.get("output_dir", "./out")
                        zip_base = os.path.join(output_dir, "cure_fixed_codebase")
                        if st.button("📦 Create ZIP of Codebase"):
                            zip_path = create_zip_archive(codebase_path, zip_base)
                            with open(zip_path, "rb") as f:
                                st.download_button(
                                    "📥 Download ZIP",
                                    data=f.read(),
                                    file_name="cure_fixed_codebase.zip",
                                    mime="application/zip",
                                )
        else:
            st.caption("No feedback data available for traceability report.")

        # ── Reset button ─────────────────────────────────────────────────
        st.divider()
        if st.button("🔄 Reset & Run Again"):
            st.session_state["fixer_in_progress"] = False
            st.session_state["fixer_complete"] = False
            st.session_state["fixer_logs"] = []
            st.session_state["fixer_phase_statuses"] = {}
            st.session_state["qa_results"] = []
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Fixer Audit Review
# ═══════════════════════════════════════════════════════════════════════════════

def _load_audit_excel(path: str) -> Dict[str, Optional[pd.DataFrame]]:
    """
    Load final_execution_audit.xlsx and return its sheets as DataFrames.

    Returns dict with keys:
        "audit_log"      — the main results sheet
        "decision_trail" — detailed audit trail (may be None)
        "summary"        — summary metadata (may be None)
    """
    result: Dict[str, Optional[pd.DataFrame]] = {
        "audit_log": None,
        "decision_trail": None,
        "summary": None,
    }
    try:
        xls = pd.ExcelFile(path, engine="openpyxl")
        for sheet in xls.sheet_names:
            lower = sheet.lower().strip()
            df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
            if "audit" in lower and "log" in lower:
                result["audit_log"] = df
            elif "decision" in lower or "trail" in lower:
                result["decision_trail"] = df
            elif "summary" in lower:
                result["summary"] = df
            else:
                # Fall back: first non-summary sheet becomes audit_log
                if result["audit_log"] is None:
                    result["audit_log"] = df
    except Exception as e:
        logger.error("Failed to load audit Excel %s: %s", path, e)
    return result


def _save_audit_feedback(
    df: pd.DataFrame,
    excel_path: str,
    sheet_name: str = "Audit Log",
) -> bool:
    """
    Write user feedback (Accepted, Feedback_Notes) columns back
    into the final_execution_audit.xlsx.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)

        ws = None
        for name in wb.sheetnames:
            if sheet_name.lower() in name.lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Build header map
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        # Ensure feedback columns exist
        next_col = ws.max_column + 1
        if "Accepted" not in header_map:
            ws.cell(row=1, column=next_col, value="Accepted")
            header_map["Accepted"] = next_col
            next_col += 1
        if "Feedback_Notes" not in header_map:
            ws.cell(row=1, column=next_col, value="Feedback_Notes")
            header_map["Feedback_Notes"] = next_col

        accepted_col = header_map["Accepted"]
        notes_col = header_map["Feedback_Notes"]

        # Build lookup
        fp_col = header_map.get("file_path")
        ln_col = header_map.get("line_number")
        it_col = header_map.get("issue_type")

        feedback_lookup = {}
        for _, row in df.iterrows():
            key = (
                str(row.get("file_path", "")).strip(),
                str(row.get("line_number", "")).strip(),
                str(row.get("issue_type", "")).strip(),
            )
            feedback_lookup[key] = (
                str(row.get("Accepted", "")).strip(),
                str(row.get("Feedback_Notes", "")).strip(),
            )

        updated = 0
        for row_idx in range(2, ws.max_row + 1):
            fp_val = str(ws.cell(row=row_idx, column=fp_col).value or "").strip() if fp_col else ""
            ln_val = str(ws.cell(row=row_idx, column=ln_col).value or "").strip() if ln_col else ""
            it_val = str(ws.cell(row=row_idx, column=it_col).value or "").strip() if it_col else ""

            key = (fp_val, ln_val, it_val)
            if key in feedback_lookup:
                accepted, notes = feedback_lookup[key]
                ws.cell(row=row_idx, column=accepted_col, value=accepted)
                ws.cell(row=row_idx, column=notes_col, value=notes)
                updated += 1

        wb.save(excel_path)
        logger.info("Updated %d rows in audit report with feedback", updated)
        return True
    except Exception as e:
        logger.error("Failed to save audit feedback: %s", e, exc_info=True)
        return False


def page_fixer_audit():
    """Review the fixer pipeline output (final_execution_audit.xlsx) and provide feedback."""
    st.markdown(
        "<div style='display:flex; align-items:center; gap:12px; margin-bottom:16px;'>"
        "<span style='font-size:36px;'>📋</span>"
        "<div>"
        "<h2 style='margin:0; padding:0;'>Fixer Audit Review</h2>"
        "<span style='color:#888; font-size:14px;'>"
        "Review fixer results, accept or reject fixes, and provide feedback"
        "</span>"
        "</div></div>",
        unsafe_allow_html=True,
    )

    out_dir = st.session_state.get("output_dir", "./out")

    # ── Section 1: Load Audit Report ──────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📂</span>"
        "<span style='font-size:18px; font-weight:600;'>1. Load Audit Report</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    # Auto-detect
    auto_path = os.path.join(out_dir, "final_execution_audit.xlsx")
    has_auto = os.path.isfile(auto_path)

    source_options = []
    if has_auto:
        source_options.append(f"Auto-detected: {os.path.basename(auto_path)}")
    source_options.append("Browse for audit Excel file")
    # Check if fixer results are in session state
    fixer_store = st.session_state.get("fixer_result_store", {})
    if fixer_store.get("audit_report_path") and os.path.isfile(str(fixer_store.get("audit_report_path", ""))):
        source_options.insert(0, "From Fix & QA tab (latest run)")

    source_choice = st.radio(
        "Select audit report source",
        source_options,
        index=0,
        key="audit_source_choice",
        label_visibility="collapsed",
    )

    audit_path = None
    if source_choice.startswith("From Fix & QA"):
        audit_path = fixer_store.get("audit_report_path")
    elif source_choice.startswith("Auto-detected"):
        audit_path = auto_path
    elif source_choice.startswith("Browse"):
        browse_path = st_tools.folder_browser(
            label="Select audit Excel file",
            default_path=out_dir,
            key="audit_excel_browser",
            show_files=True,
            file_extensions=[".xlsx", ".xls"],
            help_text="Browse for a final_execution_audit.xlsx file.",
        )
        if browse_path and os.path.isfile(browse_path):
            audit_path = browse_path

    if not audit_path or not os.path.isfile(str(audit_path)):
        st.info(
            "No audit report found. Run the fixer pipeline on the **Fix & QA** tab first, "
            "or browse for an existing `final_execution_audit.xlsx` file."
        )
        return

    st.success(f"Loaded: `{audit_path}`")

    # Load the Excel
    sheets = _load_audit_excel(audit_path)
    audit_df = sheets.get("audit_log")
    trail_df = sheets.get("decision_trail")
    summary_df = sheets.get("summary")

    if audit_df is None or audit_df.empty:
        st.warning("The audit report is empty or could not be parsed.")
        return

    st.divider()

    # ── Section 2: Summary ────────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📊</span>"
        "<span style='font-size:18px; font-weight:600;'>2. Execution Summary</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    # Compute stats from audit_df
    total = len(audit_df)
    status_col = "final_status" if "final_status" in audit_df.columns else None

    if status_col:
        fixed = (audit_df[status_col].astype(str).str.upper() == "FIXED").sum()
        failed = audit_df[status_col].astype(str).str.upper().str.contains("FAIL").sum()
        skipped = (audit_df[status_col].astype(str).str.upper() == "SKIPPED").sum()
        other = total - fixed - failed - skipped

        m1, m2, m3, m4 = st.columns(4)
        with m1:
            st.metric("Total Tasks", total)
        with m2:
            st.metric("Fixed", fixed)
        with m3:
            st.metric("Failed", failed)
        with m4:
            st.metric("Skipped", skipped)

        # Status breakdown chart
        if total > 0:
            status_counts = audit_df[status_col].astype(str).value_counts()
            st.bar_chart(status_counts, width="stretch", height=200)
    else:
        st.metric("Total Entries", total)

    # Show summary sheet if present
    if summary_df is not None and not summary_df.empty:
        with st.expander("Raw Summary Metadata", expanded=False):
            st.dataframe(summary_df, width="stretch", hide_index=True)

    st.divider()

    # ── Section 3: Audit Log with Feedback ────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📝</span>"
        "<span style='font-size:18px; font-weight:600;'>3. Audit Log — Review & Feedback</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "Review each fix result. Set **Accepted** to indicate whether you approve "
        "the fix, and add **Feedback_Notes** for any corrections or comments."
    )

    # Add feedback columns if not present
    if "Accepted" not in audit_df.columns:
        audit_df["Accepted"] = "Yes"
    if "Feedback_Notes" not in audit_df.columns:
        audit_df["Feedback_Notes"] = ""

    # Fill NaN
    audit_df["Accepted"] = audit_df["Accepted"].fillna("Yes")
    audit_df["Feedback_Notes"] = audit_df["Feedback_Notes"].fillna("")

    # Configure column display
    column_config = {
        "Accepted": st.column_config.SelectboxColumn(
            "Accepted",
            options=["Yes", "No", "Partial"],
            default="Yes",
            help="Accept, reject, or partially accept this fix.",
            width="small",
        ),
        "Feedback_Notes": st.column_config.TextColumn(
            "Feedback Notes",
            help="Comments on the fix quality, issues, or suggestions.",
            width="medium",
        ),
    }

    # Color-code status
    if status_col and status_col in audit_df.columns:
        column_config[status_col] = st.column_config.TextColumn(
            "Status",
            width="small",
        )

    # Severity filter
    filter_col1, filter_col2 = st.columns(2)
    display_df = audit_df.copy()

    with filter_col1:
        if status_col and status_col in audit_df.columns:
            statuses = ["All"] + sorted(audit_df[status_col].astype(str).unique().tolist())
            status_filter = st.selectbox("Filter by Status", statuses, key="audit_status_filter")
            if status_filter != "All":
                display_df = display_df[display_df[status_col].astype(str) == status_filter]

    with filter_col2:
        if "severity" in audit_df.columns:
            severities = ["All"] + sorted(audit_df["severity"].astype(str).unique().tolist())
            sev_filter = st.selectbox("Filter by Severity", severities, key="audit_severity_filter")
            if sev_filter != "All":
                display_df = display_df[display_df["severity"].astype(str) == sev_filter]

    st.caption(f"Showing **{len(display_df)}** of **{len(audit_df)}** entries")

    # Editable data editor
    edited_df = st.data_editor(
        display_df,
        column_config=column_config,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key="audit_editor",
    )

    # Merge edits back into the full DataFrame (in case filters are active)
    if edited_df is not None and not edited_df.equals(display_df):
        # Update only the feedback columns
        for idx in edited_df.index:
            if idx in audit_df.index:
                audit_df.at[idx, "Accepted"] = edited_df.at[idx, "Accepted"]
                audit_df.at[idx, "Feedback_Notes"] = edited_df.at[idx, "Feedback_Notes"]

    st.divider()

    # ── Section 4: Decision Trail (read-only) ─────────────────────────────
    if trail_df is not None and not trail_df.empty:
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>🔍</span>"
            "<span style='font-size:18px; font-weight:600;'>4. Decision Trail</span>"
            "</div>",
            unsafe_allow_html=True,
        )
        st.caption("Detailed audit trail of every decision made by the fixer agent.")

        with st.expander(f"Decision Trail ({len(trail_df)} entries)", expanded=False):
            st.dataframe(trail_df, width="stretch", hide_index=True)

        st.divider()

    # ── Section 5: Save & Export ──────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>💾</span>"
        "<span style='font-size:18px; font-weight:600;'>"
        f"{'5' if trail_df is not None else '4'}. Save Feedback & Export"
        "</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    save_col1, save_col2, save_col3, save_col4 = st.columns(4)

    with save_col1:
        if st.button("💾 Save Feedback to Excel", type="primary", key="audit_save_btn"):
            ok = _save_audit_feedback(audit_df, audit_path)
            if ok:
                st.session_state["_audit_feedback_saved"] = True
            else:
                st.session_state["_audit_feedback_saved"] = False

    with save_col2:
        # Download full audit with feedback
        if FEEDBACK_HELPERS_AVAILABLE:
            audit_bytes = export_to_excel_bytes(audit_df, sheet_name="Audit Log")
            st.download_button(
                "📥 Download Audit (Excel)",
                data=audit_bytes,
                file_name="fixer_audit_with_feedback.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    with save_col3:
        # Export rejected items for re-analysis
        rejected = audit_df[audit_df["Accepted"].astype(str).str.strip() == "No"]
        if not rejected.empty:
            rej_bytes = export_to_excel_bytes(rejected, sheet_name="Rejected Fixes")
            st.download_button(
                f"📥 Rejected Fixes ({len(rejected)})",
                data=rej_bytes,
                file_name="rejected_fixes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    with save_col4:
        # Feedback summary stats
        accepted_count = (audit_df["Accepted"].astype(str).str.strip() == "Yes").sum()
        rejected_count = (audit_df["Accepted"].astype(str).str.strip() == "No").sum()
        partial_count = (audit_df["Accepted"].astype(str).str.strip() == "Partial").sum()
        st.caption(
            f"Accepted: **{accepted_count}** · "
            f"Rejected: **{rejected_count}** · "
            f"Partial: **{partial_count}**"
        )

    if st.session_state.get("_audit_feedback_saved") is True:
        st.success(f"Feedback saved to `{audit_path}`")
    elif st.session_state.get("_audit_feedback_saved") is False:
        st.error("Failed to save feedback. Check logs.")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Constraints Generator
# ═══════════════════════════════════════════════════════════════════════════════

# Paths to the template / prompt files (relative to project root)
_CONSTRAINTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "agents", "constraints")
_TEMPLATE_PATH = os.path.join(_CONSTRAINTS_DIR, "TEMPLATE_constraints.md")
_PROMPT_PATH = os.path.join(_CONSTRAINTS_DIR, "GENERATE_CONSTRAINTS_PROMPT.md")


def _load_text_file(path: str) -> str:
    """Load a text file, returning empty string on failure."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return ""


def _build_constraint_generation_prompt(
    source_filename: str,
    source_code: str,
    issues_to_ignore: str,
    fix_guidelines: str,
    existing_issues_text: str,
    template_text: str,
    generator_prompt_text: str,
) -> str:
    """
    Build the full LLM prompt that combines:
    - The GENERATE_CONSTRAINTS_PROMPT instructions
    - The TEMPLATE for output format reference
    - User-supplied source code, false-positive list, and fix guidelines
    - Issues from the Excel (if provided)
    """
    # Extract just the "PROMPT" section from GENERATE_CONSTRAINTS_PROMPT.md
    # (everything after "## PROMPT (copy everything below this line)")
    prompt_section = generator_prompt_text
    marker = "## PROMPT (copy everything below this line)"
    idx = generator_prompt_text.find(marker)
    if idx >= 0:
        prompt_section = generator_prompt_text[idx + len(marker):].strip()

    parts = [prompt_section]

    parts.append(f"\n### REFERENCE TEMPLATE\nUse this exact structure for the output:\n\n```markdown\n{template_text}\n```")

    parts.append(f"\n### INPUT\n\n**Source file**: `{source_filename}`\n")

    if existing_issues_text:
        parts.append(
            f"**Issues from analysis report** (these are issues the tool flagged — "
            f"review them and convert false positives into IGNORE rules, "
            f"and genuine issues into Resolution Rules):\n\n{existing_issues_text}\n"
        )

    if issues_to_ignore.strip():
        parts.append(f"**Issues to ignore (false positives)**:\n{issues_to_ignore.strip()}\n")

    if fix_guidelines.strip():
        parts.append(f"**Fix guidelines**:\n{fix_guidelines.strip()}\n")

    if source_code.strip():
        parts.append(f"**Source code**:\n```\n{source_code.strip()}\n```\n")

    parts.append("\nNow generate the constraints file. Output ONLY the Markdown content, no extra commentary.")

    return "\n\n".join(parts)


def _format_excel_issues_for_prompt(df: pd.DataFrame, target_file: str = "") -> str:
    """
    Format issues from a DataFrame (detailed_code_review.xlsx format)
    into a text block suitable for the LLM prompt.

    If target_file is specified, only include issues for that file.
    """
    if df is None or df.empty:
        return ""

    # Filter to target file if specified
    if target_file:
        stem = Path(target_file).stem
        mask = df["File"].astype(str).apply(
            lambda f: stem in f or target_file in f
        )
        df = df[mask]

    if df.empty:
        return ""

    lines = []
    for _, row in df.iterrows():
        severity = str(row.get("Severity", "")).strip()
        title = str(row.get("Title", "")).strip()
        line_no = str(row.get("Line", "")).strip()
        category = str(row.get("Category", "")).strip()
        desc = str(row.get("Description", "")).strip()
        suggestion = str(row.get("Suggestion", "")).strip()
        code = str(row.get("Code", "")).strip()
        action = str(row.get("Action", row.get("Feedback", ""))).strip()

        entry = f"- [{severity}] {title} (Line {line_no}, {category}): {desc}"
        if suggestion:
            entry += f"\n  Suggestion: {suggestion}"
        if code:
            entry += f"\n  Code: `{code[:200]}`"
        if action and action.lower() in ("skip", "review"):
            entry += f"\n  User action: {action}"
        lines.append(entry)

    return "\n".join(lines)


def page_constraints_generator():
    """UI page for generating <filename>_constraints.md files using LLM."""
    st.markdown(
        "<div style='display:flex; align-items:center; gap:12px; margin-bottom:16px;'>"
        "<span style='font-size:36px;'>📐</span>"
        "<div>"
        "<h2 style='margin:0; padding:0;'>Constraints Generator</h2>"
        "<span style='color:#888; font-size:14px;'>"
        "Create constraint files to suppress false positives and guide code fixes"
        "</span>"
        "</div></div>",
        unsafe_allow_html=True,
    )

    # Load template and prompt
    template_text = _load_text_file(_TEMPLATE_PATH)
    generator_prompt_text = _load_text_file(_PROMPT_PATH)

    if not template_text:
        st.warning(f"Template file not found: `{_TEMPLATE_PATH}`")
    if not generator_prompt_text:
        st.warning(f"Generator prompt not found: `{_PROMPT_PATH}`")

    # ── Codebase-Wide Auto-Generation ─────────────────────────────────────
    with st.expander("🔬 Auto-Generate Codebase Constraints (scans all symbols)", expanded=False):
        st.caption(
            "Scans the entire codebase for enums, structs, macros, bit fields, "
            "and helper functions, then generates a `codebase_constraints.md` "
            "file with IGNORE rules for common false positives."
        )
        codebase_path_for_gen = st.session_state.get("codebase_path", "")
        if codebase_path_for_gen and os.path.isdir(codebase_path_for_gen):
            if st.button("🔬 Generate Codebase Constraints", key="auto_gen_codebase_btn"):
                with st.spinner("Scanning codebase for symbols..."):
                    try:
                        from agents.context.codebase_constraint_generator import generate_constraints
                        md_text = generate_constraints(
                            codebase_path=codebase_path_for_gen,
                            exclude_dirs=[d.strip() for d in st.session_state.get("exclude_dirs", "").split(",") if d.strip()],
                            exclude_globs=[g.strip() for g in st.session_state.get("exclude_globs", "").split(",") if g.strip()],
                        )
                        # Save to constraints folder
                        save_path = os.path.join(_CONSTRAINTS_DIR, "codebase_constraints.md")
                        os.makedirs(_CONSTRAINTS_DIR, exist_ok=True)
                        with open(save_path, "w", encoding="utf-8") as f:
                            f.write(md_text)
                        st.success(f"Generated `codebase_constraints.md` ({md_text.count(chr(10)) + 1} lines)")
                        st.download_button(
                            "📥 Download codebase_constraints.md",
                            data=md_text,
                            file_name="codebase_constraints.md",
                            mime="text/markdown",
                            key="dl_codebase_constraints",
                        )
                    except Exception as e:
                        st.error(f"Generation failed: {e}")
                        logger.error("Codebase constraint generation failed", exc_info=True)
        else:
            st.info("Set a codebase path in the Analyze tab first.")

    st.divider()

    # ── Section 1: Source File ────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📄</span>"
        "<span style='font-size:18px; font-weight:600;'>1. Source File</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    source_file_path = st_tools.folder_browser(
        label="Source File Path",
        default_path=st.session_state.get("codebase_path", ""),
        key="constraints_source_browser",
        show_files=True,
        file_extensions=[".v", ".sv", ".svh", ".vh"],
        help_text="Path to the source file you want to create constraints for.",
    )

    source_code = ""
    source_filename = ""
    if source_file_path and os.path.isfile(source_file_path):
        source_filename = os.path.basename(source_file_path)
        try:
            with open(source_file_path, "r", encoding="utf-8", errors="replace") as f:
                source_code = f.read()
            st.success(f"Loaded `{source_filename}` ({len(source_code):,} chars)")
        except Exception as e:
            st.error(f"Failed to read file: {e}")
    elif source_file_path:
        # User may have typed just a filename
        source_filename = os.path.basename(source_file_path)
        st.info(f"File not found at path — you can still type issues manually. Target: `{source_filename}`")

    st.divider()

    # ── Section 2: Issues Excel (optional) ────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>📊</span>"
        "<span style='font-size:18px; font-weight:600;'>2. Issues from Analysis (optional)</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "Load issues from a `detailed_code_review.xlsx` file. "
        "The LLM will analyze these issues and convert false positives into IGNORE rules."
    )

    issues_df = None
    out_dir = st.session_state.get("output_dir", "./out")

    issues_source = st.radio(
        "Issue source",
        [
            "Auto-detect from output folder",
            "Browse for Excel file",
            "Use issues from Review tab",
            "Skip (manual entry only)",
        ],
        horizontal=True,
        key="constraints_issues_source",
        label_visibility="collapsed",
    )

    if issues_source == "Auto-detect from output folder":
        auto_path = os.path.join(out_dir, "detailed_code_review.xlsx")
        if os.path.isfile(auto_path) and FEEDBACK_HELPERS_AVAILABLE:
            issues_df = load_excel_as_dataframe(auto_path)
            if issues_df is not None:
                st.success(f"Loaded `{auto_path}` — **{len(issues_df)}** issues")
            else:
                st.warning("Could not parse the Excel file.")
        else:
            st.info("No `detailed_code_review.xlsx` found in output folder.")

    elif issues_source == "Browse for Excel file":
        excel_path = st_tools.folder_browser(
            label="Select Excel file",
            default_path=out_dir,
            key="constraints_excel_browser",
            show_files=True,
            file_extensions=[".xlsx", ".xls"],
            help_text="Browse for an analysis Excel file.",
        )
        if excel_path and os.path.isfile(excel_path) and FEEDBACK_HELPERS_AVAILABLE:
            issues_df = load_excel_as_dataframe(excel_path)
            if issues_df is not None:
                st.success(f"Loaded — **{len(issues_df)}** issues")

    elif issues_source == "Use issues from Review tab":
        review_df = st.session_state.get("feedback_df")
        if review_df is not None:
            issues_df = review_df
            st.success(f"Using **{len(issues_df)}** issues from the Review tab")
        else:
            st.info("No issues available from the Review tab. Run analysis first.")

    # Show preview pane if issues loaded
    if issues_df is not None and not issues_df.empty:
        preview_cols = [c for c in ["File", "Title", "Severity", "Category", "Line", "Description", "Suggestion"] if c in issues_df.columns]
        preview_df = issues_df[preview_cols] if preview_cols else issues_df

        # Summary metrics row
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-top:8px; margin-bottom:4px;'>"
            "<span style='font-size:16px;'>🔎</span>"
            "<span style='font-size:15px; font-weight:600;'>Issues Preview</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        if "Severity" in issues_df.columns:
            sev_counts = issues_df["Severity"].astype(str).str.strip().str.title().value_counts()
            metric_cols = st.columns(min(len(sev_counts) + 1, 6))
            metric_cols[0].metric("Total", len(issues_df))
            for i, (sev, count) in enumerate(sev_counts.items()):
                if i + 1 < len(metric_cols):
                    metric_cols[i + 1].metric(sev, count)

        # File-filtered view if a source file is selected
        if source_filename and "File" in issues_df.columns:
            stem = Path(source_filename).stem
            file_mask = issues_df["File"].astype(str).apply(lambda f: stem in f)
            file_issues = issues_df[file_mask]
            file_preview_cols = [c for c in preview_cols if c != "File"]

            if not file_issues.empty:
                with st.expander(f"Issues for `{source_filename}` ({len(file_issues)} found)", expanded=True):
                    st.dataframe(
                        file_issues[file_preview_cols].head(50) if file_preview_cols else file_issues.head(50),
                        width="stretch",
                        hide_index=True,
                    )
            else:
                st.info(f"No issues found specifically for `{source_filename}`. All issues will be included as context.")

        # Full issues table
        with st.expander(f"All Issues ({len(issues_df)} total)", expanded=False):
            st.dataframe(
                preview_df.head(100),
                width="stretch",
                hide_index=True,
            )

    st.divider()

    # ── Section 3: Manual Inputs ──────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>✏️</span>"
        "<span style='font-size:18px; font-weight:600;'>3. Manual Inputs</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    issues_to_ignore = st.text_area(
        "Issues to Ignore (false positives)",
        height=120,
        placeholder=(
            "One per line. Example:\n"
            "- IGNORE NULL check for `vdev` in dp_rx_process() — validated at entry point\n"
            "- IGNORE bounds check for `queue_id` — hardware-limited to MAX_QUEUES\n"
            "- IGNORE unused parameter `cookie` — required by kernel callback signature"
        ),
        key="constraints_ignore_list",
    )

    fix_guidelines = st.text_area(
        "Fix Guidelines (optional)",
        height=100,
        placeholder=(
            "Example:\n"
            "- Don't add locks in ISR paths\n"
            "- Use kernel error codes (-EINVAL, -ENOMEM), not custom enums\n"
            "- Prefer devm_kzalloc for configuration data"
        ),
        key="constraints_fix_guidelines",
    )

    include_source = st.checkbox(
        "Include source code in prompt (helps LLM infer additional constraints)",
        value=bool(source_code),
        key="constraints_include_source",
    )

    st.divider()

    # ── Section 4: Generate ───────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
        "<span style='font-size:20px;'>🤖</span>"
        "<span style='font-size:18px; font-weight:600;'>4. Generate Constraints File</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    if not source_filename:
        st.warning("Select or enter a source file path above to generate constraints.")
        return

    target_stem = Path(source_filename).stem
    output_name = f"{target_stem}_constraints.md"
    st.info(f"Will generate: **`{output_name}`**")

    gen_col1, gen_col2 = st.columns([1, 3])

    with gen_col1:
        generate_clicked = st.button("🤖 Generate with LLM", type="primary", key="constraints_generate_btn")

    with gen_col2:
        use_template = st.button("📝 Start from Template", key="constraints_template_btn")

    # ── Generate from template (no LLM) ──────────────────────────────────
    if use_template:
        # Pre-fill the template with the filename
        filled = template_text.replace("<FILENAME>", source_filename)
        filled = filled.replace(
            "# NAMING CONVENTION:",
            f"# Generated for: {source_filename}\n# NAMING CONVENTION:",
        )
        st.session_state["constraints_generated_md"] = filled
        # Clear the editor widget key so it picks up the new value
        st.session_state.pop("constraints_editor", None)
        st.rerun()

    # ── Generate with LLM ────────────────────────────────────────────────
    if generate_clicked:
        # Build the issues text from Excel
        existing_issues_text = ""
        if issues_df is not None:
            existing_issues_text = _format_excel_issues_for_prompt(issues_df, source_filename)

        # Build the full prompt
        full_prompt = _build_constraint_generation_prompt(
            source_filename=source_filename,
            source_code=source_code if include_source else "",
            issues_to_ignore=issues_to_ignore,
            fix_guidelines=fix_guidelines,
            existing_issues_text=existing_issues_text,
            template_text=template_text,
            generator_prompt_text=generator_prompt_text,
        )

        # Call the LLM
        with st.spinner("Generating constraints with LLM..."):
            try:
                from utils.common.llm_tools import LLMTools
                llm = LLMTools()
                response = llm.llm_call(full_prompt)

                # Clean up response — strip markdown code fences if the LLM wrapped it
                cleaned = response.strip()
                if cleaned.startswith("```markdown"):
                    cleaned = cleaned[len("```markdown"):].strip()
                elif cleaned.startswith("```md"):
                    cleaned = cleaned[len("```md"):].strip()
                elif cleaned.startswith("```"):
                    cleaned = cleaned[3:].strip()
                if cleaned.endswith("```"):
                    cleaned = cleaned[:-3].strip()

                st.session_state["constraints_generated_md"] = cleaned
                # Clear the editor widget key so it picks up the new value
                st.session_state.pop("constraints_editor", None)
                st.rerun()

            except Exception as e:
                st.error(f"LLM call failed: {e}")
                logger.error("Constraint generation LLM call failed", exc_info=True)

    # ── Section 5: Preview & Edit ─────────────────────────────────────────
    generated_md = st.session_state.get("constraints_generated_md", "")
    if generated_md:
        st.divider()

        # Auto-save the generated file to the output directory for easy access
        auto_save_dir = st.session_state.get("output_dir", "./out")
        auto_save_path = os.path.join(auto_save_dir, output_name)
        try:
            os.makedirs(auto_save_dir, exist_ok=True)
            with open(auto_save_path, "w", encoding="utf-8") as f:
                f.write(generated_md)
        except Exception:
            auto_save_path = None

        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>👁️</span>"
            "<span style='font-size:18px; font-weight:600;'>5. Preview & Edit</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        if auto_save_path:
            st.success(f"Constraints file generated: `{auto_save_path}`")

        # Editable text area with the generated content
        edited_md = st.text_area(
            "Edit the generated constraints file",
            value=generated_md,
            height=500,
            key="constraints_editor",
        )

        # Preview as rendered markdown
        with st.expander("Preview (rendered Markdown)", expanded=False):
            st.markdown(edited_md, unsafe_allow_html=True)

        st.divider()

        # ── Section 6: Save & Download ────────────────────────────────────
        st.markdown(
            "<div style='display:flex; align-items:center; gap:8px; margin-bottom:8px;'>"
            "<span style='font-size:20px;'>💾</span>"
            "<span style='font-size:18px; font-weight:600;'>6. Save & Download</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        dl_col, save_col, clear_col = st.columns(3)

        with dl_col:
            # Download .md file directly from the page
            st.download_button(
                "📥 Download .md File",
                data=edited_md,
                file_name=output_name,
                mime="text/markdown",
                type="primary",
                key="constraints_download_btn",
            )

        with save_col:
            # Save directly to agents/constraints/
            if st.button("💾 Save to Constraints Folder", key="constraints_save_btn"):
                save_path = os.path.join(_CONSTRAINTS_DIR, output_name)
                try:
                    os.makedirs(_CONSTRAINTS_DIR, exist_ok=True)
                    with open(save_path, "w", encoding="utf-8") as f:
                        f.write(edited_md)
                    st.success(f"Saved to `{save_path}`")
                except Exception as e:
                    st.error(f"Failed to save: {e}")

        with clear_col:
            # Clear / reset
            if st.button("🗑️ Clear", key="constraints_clear_btn"):
                st.session_state["constraints_generated_md"] = ""
                st.session_state.pop("constraints_editor", None)
                st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Telemetry Dashboard
# ═══════════════════════════════════════════════════════════════════════════════

def page_telemetry():
    """Enhanced telemetry dashboard with tabbed layout for deep analytics."""
    st.markdown(
        "<h2 style='text-align:center; margin-top:-10px;'>"
        "📈 Telemetry Dashboard</h2>",
        unsafe_allow_html=True,
    )

    # Try to load TelemetryService
    telemetry = None
    try:
        from db.telemetry_service import TelemetryService
        from utils.parsers.global_config_parser import GlobalConfig
        gc = GlobalConfig()
        conn_str = gc.get("POSTGRES_CONNECTION")
        if conn_str:
            db_cfg = gc.get("database", {}) or {}
            telemetry = TelemetryService(
                connection_string=conn_str,
                pool_size=int(db_cfg.get("pool_size", 5)),
                pool_recycle=int(db_cfg.get("pool_recycle", 3600)),
                pool_timeout=int(db_cfg.get("pool_timeout", 30)),
                pool_pre_ping=bool(db_cfg.get("pool_pre_ping", True)),
            )
    except Exception:
        pass

    if telemetry is None or not telemetry.enabled:
        st.info(
            "Telemetry is not available. Ensure PostgreSQL is configured "
            "and `telemetry.enable: true` is set in global_config.yaml."
        )
        return

    # ── Summary metrics (always visible above tabs) ──────────────────────
    stats = telemetry.get_summary_stats()
    if not stats:
        st.info("No telemetry data recorded yet. Run an analysis to start collecting metrics.")
        return

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Runs", stats.get("total_runs", 0))
    m2.metric("Total Issues Found", stats.get("total_issues", 0))
    m3.metric("Total Issues Fixed", stats.get("total_fixed", 0))
    m4.metric("Fix Success Rate", f"{stats.get('fix_success_rate', 0)}%")

    # ── Cost summary (top-level) ─────────────────────────────────────────
    cost_data = telemetry.get_cost_summary(days=30)
    cost_total = 0.0
    if cost_data and cost_data.get("total_cost") is not None:
        cost_total = float(cost_data["total_cost"])
    m5, m6, m7, m8 = st.columns(4)
    m5.metric("Analysis Runs", stats.get("analysis_runs", 0))
    m6.metric("Fixer Runs", stats.get("fixer_runs", 0))
    m7.metric("Patch Runs", stats.get("patch_runs", 0))
    m8.metric("Est. Cost (30d)", f"${cost_total:,.2f}")

    st.markdown("---")

    # ═════════════════════════════════════════════════════════════════════
    #  Tabbed layout
    # ═════════════════════════════════════════════════════════════════════
    tab_overview, tab_findings, tab_llm, tab_constraints, tab_reports = st.tabs([
        "📊 Overview", "🔍 Detailed Findings", "🤖 LLM Analytics",
        "🛡️ Constraints & Quality", "📋 Usage Reports",
    ])

    # ── Tab 1: Overview ─────────────────────────────────────────────────
    with tab_overview:
        # Runs over time
        runs_by_date = stats.get("runs_by_date", {})
        if runs_by_date:
            st.markdown("#### Runs Over Time (Last 30 Days)")
            chart_df = pd.DataFrame(
                list(runs_by_date.items()), columns=["Date", "Runs"],
            )
            chart_df["Date"] = pd.to_datetime(chart_df["Date"])
            st.bar_chart(chart_df.set_index("Date"))

        # Cost trend
        if cost_data and cost_data.get("daily_trend"):
            st.markdown("#### Daily Cost Trend (Last 30 Days)")
            trend = cost_data["daily_trend"]
            trend_df = pd.DataFrame(trend)
            if not trend_df.empty and "date" in trend_df.columns and "cost" in trend_df.columns:
                trend_df["date"] = pd.to_datetime(trend_df["date"])
                trend_df = trend_df.rename(columns={"date": "Date", "cost": "Cost ($)"})
                st.line_chart(trend_df.set_index("Date"))

        # Severity + issue types side by side
        col_sev, col_types = st.columns(2)
        with col_sev:
            issues_by_sev = stats.get("issues_by_severity", {})
            if issues_by_sev:
                st.markdown("#### Issues by Severity")
                sev_df = pd.DataFrame(
                    list(issues_by_sev.items()), columns=["Severity", "Count"],
                )
                st.bar_chart(sev_df.set_index("Severity"))
        with col_types:
            top_types = stats.get("top_issue_types", {})
            if top_types:
                st.markdown("#### Top Issue Types")
                types_df = pd.DataFrame(
                    list(top_types.items()), columns=["Issue Type", "Count"],
                )
                st.dataframe(types_df, use_container_width=True, hide_index=True)

        # LLM usage summary
        st.markdown("#### LLM Usage Summary")
        llm_col1, llm_col2 = st.columns(2)
        with llm_col1:
            st.metric("Total LLM Calls", stats.get("total_llm_calls", 0))
            st.metric("Total Prompt Tokens", f"{stats.get('total_prompt_tokens', 0):,}")
        with llm_col2:
            st.metric("Total Completion Tokens", f"{stats.get('total_completion_tokens', 0):,}")
            st.metric("Avg Duration", f"{stats.get('avg_duration', 0):.1f}s")

        # Recent runs table
        st.markdown("#### Recent Runs")
        recent = telemetry.get_recent_runs(limit=25)
        if recent:
            display_cols = [
                "run_id", "created_at", "mode", "status",
                "files_analyzed", "issues_total", "issues_fixed",
                "issues_skipped", "issues_failed", "duration_seconds",
                "llm_model", "use_verible",
            ]
            runs_df = pd.DataFrame(recent)
            available = [c for c in display_cols if c in runs_df.columns]
            st.dataframe(runs_df[available], use_container_width=True, hide_index=True)

            # Drill-down into a run
            run_ids = [r["run_id"] for r in recent]
            selected_run = st.selectbox(
                "View events for run:", ["(select)"] + run_ids, key="overview_run_select",
            )
            if selected_run and selected_run != "(select)":
                events = telemetry.get_run_events(selected_run)
                if events:
                    ev_df = pd.DataFrame(events)
                    ev_cols = [c for c in [
                        "created_at", "event_type", "file_path",
                        "issue_type", "severity", "llm_model",
                        "prompt_tokens", "completion_tokens", "latency_ms",
                    ] if c in ev_df.columns]
                    st.dataframe(ev_df[ev_cols], use_container_width=True, hide_index=True)
                else:
                    st.info("No events recorded for this run.")
        else:
            st.info("No runs recorded yet.")

    # ── Tab 2: Detailed Findings ────────────────────────────────────────
    with tab_findings:
        st.markdown("#### Granular Finding Explorer")

        # False positive rate metric
        fp_data = telemetry.get_false_positive_rate(days=30)
        fp_rate = 0.0
        fp_total = 0
        if fp_data:
            fp_rate = float(fp_data.get("false_positive_rate", 0))
            fp_total = int(fp_data.get("total_findings", 0))
        fp_c1, fp_c2, fp_c3 = st.columns(3)
        fp_c1.metric("Findings (30d)", fp_total)
        fp_c2.metric("False Positive Rate", f"{fp_rate:.1f}%")
        fp_c3.metric("Confirmed Findings", int(fp_data.get("confirmed", 0)) if fp_data else 0)

        # Filter by run
        recent_for_filter = telemetry.get_recent_runs(limit=50)
        run_opts = ["All Runs"]
        if recent_for_filter:
            run_opts += [r["run_id"] for r in recent_for_filter]
        filter_run = st.selectbox(
            "Filter by Run ID:", run_opts, key="findings_run_filter",
        )

        sel_run_id = None if filter_run == "All Runs" else filter_run
        findings = telemetry.get_findings_detail(run_id=sel_run_id)

        if findings:
            findings_df = pd.DataFrame(findings)
            # Severity and category filters
            f_col1, f_col2 = st.columns(2)
            with f_col1:
                if "severity" in findings_df.columns:
                    sev_opts = ["All"] + sorted(findings_df["severity"].dropna().unique().tolist())
                    sel_sev = st.selectbox("Filter Severity:", sev_opts, key="findings_sev_filter")
                    if sel_sev != "All":
                        findings_df = findings_df[findings_df["severity"] == sel_sev]
            with f_col2:
                if "category" in findings_df.columns:
                    cat_opts = ["All"] + sorted(findings_df["category"].dropna().unique().tolist())
                    sel_cat = st.selectbox("Filter Category:", cat_opts, key="findings_cat_filter")
                    if sel_cat != "All":
                        findings_df = findings_df[findings_df["category"] == sel_cat]

            show_cols = [c for c in [
                "finding_id", "run_id", "created_at", "file_path",
                "line_start", "title", "category", "severity",
                "confidence", "description", "is_false_positive",
            ] if c in findings_df.columns]
            st.dataframe(findings_df[show_cols], use_container_width=True, hide_index=True)

            # CSV export
            csv = findings_df[show_cols].to_csv(index=False)
            st.download_button(
                "📥 Export Findings CSV", csv,
                file_name="telemetry_findings.csv", mime="text/csv",
                key="findings_csv_dl",
            )
        else:
            st.info("No findings recorded yet.")

    # ── Tab 3: LLM Analytics ────────────────────────────────────────────
    with tab_llm:
        st.markdown("#### LLM Call Analytics")

        # Cost by provider/model
        if cost_data and cost_data.get("by_model"):
            st.markdown("##### Cost by Provider / Model")
            model_cost_df = pd.DataFrame(cost_data["by_model"])
            if not model_cost_df.empty:
                mc_c1, mc_c2 = st.columns([2, 1])
                with mc_c1:
                    if "model" in model_cost_df.columns and "cost" in model_cost_df.columns:
                        chart_mc = model_cost_df.rename(columns={"model": "Model", "cost": "Cost ($)"})
                        st.bar_chart(chart_mc.set_index("Model"))
                with mc_c2:
                    st.dataframe(model_cost_df, use_container_width=True, hide_index=True)

        # Token efficiency
        st.markdown("##### Token Breakdown by Model")
        llm_usage = telemetry.get_llm_usage_stats()
        by_model = llm_usage.get("by_model", [])
        if by_model:
            llm_df = pd.DataFrame(by_model)
            st.dataframe(llm_df, use_container_width=True, hide_index=True)

        # Per-call detail for a specific run
        st.markdown("##### Per-Call Detail")
        recent_llm = telemetry.get_recent_runs(limit=30)
        llm_run_opts = ["(select run)"]
        if recent_llm:
            llm_run_opts += [r["run_id"] for r in recent_llm]
        llm_sel_run = st.selectbox(
            "View LLM calls for run:", llm_run_opts, key="llm_run_select",
        )
        if llm_sel_run and llm_sel_run != "(select run)":
            try:
                with telemetry._engine.connect() as conn:
                    from sqlalchemy import text as sa_text
                    rows = conn.execute(sa_text(
                        "SELECT call_id, created_at, provider, model, purpose, "
                        "file_path, chunk_index, prompt_tokens, completion_tokens, "
                        "total_tokens, latency_ms, estimated_cost_usd, status "
                        "FROM telemetry_llm_calls WHERE run_id = :rid "
                        "ORDER BY created_at"
                    ), {"rid": llm_sel_run}).fetchall()
                if rows:
                    cols = [
                        "call_id", "created_at", "provider", "model", "purpose",
                        "file_path", "chunk_index", "prompt_tokens", "completion_tokens",
                        "total_tokens", "latency_ms", "estimated_cost_usd", "status",
                    ]
                    calls_df = pd.DataFrame(rows, columns=cols)
                    st.dataframe(calls_df, use_container_width=True, hide_index=True)

                    # Latency distribution
                    if "latency_ms" in calls_df.columns:
                        st.markdown("##### Latency Distribution (ms)")
                        st.bar_chart(calls_df["latency_ms"].value_counts().sort_index())

                    # Summary for run
                    sc1, sc2, sc3 = st.columns(3)
                    sc1.metric("Calls", len(calls_df))
                    sc2.metric("Total Tokens", f"{calls_df['total_tokens'].sum():,}")
                    run_cost = calls_df["estimated_cost_usd"].sum()
                    sc3.metric("Run Cost", f"${float(run_cost):,.4f}")
                else:
                    st.info("No LLM call records for this run.")
            except Exception as e:
                st.warning(f"Could not load LLM call details: {e}")

    # ── Tab 4: Constraints & Quality ────────────────────────────────────
    with tab_constraints:
        st.markdown("#### Constraint Effectiveness & Quality")

        # Constraint hits summary
        constraint_data = telemetry.get_constraint_effectiveness()
        if constraint_data and constraint_data.get("by_rule"):
            st.markdown("##### Constraint Hit Summary")
            cr_df = pd.DataFrame(constraint_data["by_rule"])
            if not cr_df.empty:
                st.dataframe(cr_df, use_container_width=True, hide_index=True)

            by_action = constraint_data.get("by_action", {})
            if by_action:
                st.markdown("##### Actions Breakdown")
                act_df = pd.DataFrame(
                    list(by_action.items()), columns=["Action", "Count"],
                )
                st.bar_chart(act_df.set_index("Action"))
        else:
            st.info("No constraint hit data recorded yet.")

        st.markdown("---")

        # Agent comparison
        st.markdown("##### Agent Comparison (Last 30 Days)")
        agent_cmp = telemetry.get_agent_comparison(days=30)
        if agent_cmp:
            cmp_df = pd.DataFrame(agent_cmp)
            if not cmp_df.empty:
                st.dataframe(cmp_df, use_container_width=True, hide_index=True)
        else:
            st.info("No agent comparison data available yet.")

        st.markdown("---")

        # False-positive rate trend
        st.markdown("##### False-Positive Rate (30d)")
        if fp_data:
            fp_detail = {
                "Total Findings": fp_data.get("total_findings", 0),
                "False Positives": fp_data.get("false_positives", 0),
                "Confirmed": fp_data.get("confirmed", 0),
                "Rate": f"{fp_data.get('false_positive_rate', 0):.1f}%",
            }
            st.json(fp_detail)
        else:
            st.info("No false-positive data available.")

    # ── Tab 5: Usage Reports ────────────────────────────────────────────
    with tab_reports:
        st.markdown("#### Usage Reports")

        rpt_col1, rpt_col2 = st.columns([1, 2])
        with rpt_col1:
            report_type = st.radio(
                "Report Type", ["daily", "weekly"], key="report_type_radio",
            )
            if st.button("🔄 Generate Today's Report", key="gen_report_btn"):
                try:
                    from datetime import date as dt_date
                    telemetry.generate_usage_report(
                        report_date=dt_date.today().isoformat(),
                        report_type=report_type,
                    )
                    st.success(f"Generated {report_type} report for today.")
                except Exception as e:
                    st.error(f"Report generation failed: {e}")

        with rpt_col2:
            reports = telemetry.get_usage_reports(report_type=report_type, limit=30)
            if reports:
                rpt_df = pd.DataFrame(reports)
                show_rpt_cols = [c for c in [
                    "report_date", "report_type", "total_runs", "total_files",
                    "total_findings", "total_fixes", "total_tokens",
                    "estimated_cost_usd",
                ] if c in rpt_df.columns]
                st.dataframe(rpt_df[show_rpt_cols], use_container_width=True, hide_index=True)

                # CSV export
                rpt_csv = rpt_df[show_rpt_cols].to_csv(index=False)
                st.download_button(
                    "📥 Export Report CSV", rpt_csv,
                    file_name=f"usage_reports_{report_type}.csv", mime="text/csv",
                    key="reports_csv_dl",
                )
            else:
                st.info(f"No {report_type} reports generated yet. Click 'Generate' to create one.")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: Chat (existing, preserved)
# ═══════════════════════════════════════════════════════════════════════════════

def page_chat():
    """Interactive codebase health chat powered by LLM orchestration."""
    st.markdown(
        "<h2 style='text-align:center; margin-top:-10px;'>"
        "Codebase Health Chat</h2>",
        unsafe_allow_html=True,
    )

    if not CHAT_AGENT_AVAILABLE or orchestrator is None:
        st.error(
            "Chat agent is not available. Ensure `agents/codebase_analysis_chat_agent.py` "
            "is installed and that the vector DB has been populated via `main.py --enable-vector-db`."
        )
        return

    # Welcome message
    with st.chat_message("assistant", avatar=APP_ICON):
        st.markdown(
            f"<b>Welcome to <span style='color:{st_tools.CARE_PRIMARY};'>CARE</span> "
            "Codebase Health Chat!</b><br>"
            "Ask about dependencies, complexity, security, documentation, "
            "maintainability, test coverage, and refactoring recommendations.",
            unsafe_allow_html=True,
        )

    st_tools.feedback_info_if_enabled()

    # Sample queries
    with st.expander("Example questions you can ask"):
        st.markdown(
            "- **Module deep-dive**: _Show all details about the auth module — "
            "dependencies, security risks, test coverage, and documentation gaps._\n"
            "- **Overall health**: _Summarize the codebase health across all dimensions. "
            "Highlight the top 3 issues I should fix first._\n"
            "- **Security audit**: _List all high-severity security findings with file "
            "names and line numbers._\n"
            "- **Dead code**: _Which functions are unreachable from any entry point?_\n"
            "- **Complexity hotspots**: _Show functions with cyclomatic complexity above 25._"
        )

    # Render existing history
    if st.session_state.chat_summary:
        st.info("Earlier conversation summary: " + st.session_state.chat_summary)

    for idx, (speaker, text) in enumerate(st.session_state.chat_history):
        role = "user" if speaker == "You" else "assistant"
        avatar = "🧑" if role == "user" else APP_ICON
        with st.chat_message(role, avatar=avatar):
            if role == "assistant":
                _render_markdown_with_tables(_extract_content(text))
            else:
                st.markdown(text)

            if role == "assistant":
                user_msg = ""
                if idx > 0 and st.session_state.chat_history[idx - 1][0] == "You":
                    user_msg = st.session_state.chat_history[idx - 1][1]
                feedback = st_tools.feedback_widget(idx, user_msg, text)
                if feedback:
                    st.session_state["all_feedback"].append(feedback)

    # Summarize old turns
    max_turns = 25
    history = st.session_state.chat_history
    if len(history) > max_turns:
        old_messages = history[:-max_turns]
        st.session_state.chat_summary = st_tools.summarize_chat(
            old_messages, st.session_state.chat_summary
        )
        st.session_state.chat_history = history[-max_turns:]

    # New user input
    user_input = st.chat_input("Ask about your codebase's health and metrics:")
    if user_input:
        st.session_state.chat_history.append(("You", user_input))
        st.session_state.chat_history.append(("Assistant", PLACEHOLDER))
        st.rerun()

    # Process pending placeholder
    if (
        st.session_state.chat_history
        and st.session_state.chat_history[-1] == ("Assistant", PLACEHOLDER)
    ):
        user_message = (
            st.session_state.chat_history[-2][1]
            if len(st.session_state.chat_history) >= 2
            else ""
        )
        with st.spinner("CARE is analyzing..."):
            try:
                state = CodebaseAnalysisSessionState(user_input=user_message)
                state = orchestrator.run_multiturn_chain(state)
                answer = _extract_content(state.formatted_response)
            except Exception as e:
                logger.error("Chat orchestration error: %s", e, exc_info=True)
                st.error(f"Error: {e}")
                answer = "Sorry, I couldn't process that request. Please try rephrasing."

        st.session_state.chat_history[-1] = ("Assistant", answer)
        st.rerun()

    # Download full history
    if st.session_state.get("chat_history"):
        st.download_button(
            "Download Chat History",
            "\n".join(
                f"{speaker}: {_extract_content(text) if speaker != 'You' else text}"
                for speaker, text in st.session_state["chat_history"]
            ),
            file_name="cure_chat_history.txt",
        )


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: About
# ═══════════════════════════════════════════════════════════════════════════════

def page_about():
    """About page with project overview and connection info."""
    col1, col2 = st.columns([1, 3])
    with col1:
        if os.path.isfile(LOGO_SIDEBAR):
            st.image(LOGO_SIDEBAR, width=120)
    with col2:
        st.markdown(
            "## About CARE\n\n"
            "**CARE** (Codebase Analysis & Repair Engine) is a multi-stage Verilog/SystemVerilog HDL codebase "
            "health analysis pipeline. It combines fast regex-based static analyzers with "
            "deep static analysis adapters (Lizard, Flawfinder, CCLS/libclang) and "
            "LLM-powered code review to produce actionable health metrics.\n\n"
            "**Key features:**\n\n"
            "- 9 built-in health analyzers (complexity, security, memory, deadlocks, etc.)\n"
            "- Deep static adapters: AST complexity, dead code detection, call graph analysis\n"
            "- Multi-provider LLM support (Anthropic, QGenie, Vertex AI, Azure OpenAI)\n"
            "- Human-in-the-loop agentic code repair\n"
            "- Vector DB ingestion for RAG-powered chat\n"
        )

    st.divider()

    # FAQ content (merged from old page_faq)
    st.markdown("### Frequently Asked Questions")
    FAQS = [
        (
            "What can I ask in the chat?",
            "Ask about code health metrics, module dependencies, security findings, "
            "documentation gaps, test coverage, complexity hotspots, dead code, "
            "and refactoring recommendations.",
        ),
        (
            "What data does this use?",
            "It uses precomputed codebase analysis reports (healthreport.json), "
            "dependency graphs, static analysis adapter results, and vector DB "
            "embeddings generated by the CARE pipeline.",
        ),
        (
            "How do I populate the data?",
            "Run the analysis pipeline first:\n\n"
            "```bash\n"
            "python main.py --codebase-path /path/to/project --enable-vector-db --enable-adapters\n"
            "```\n\n"
            "Or use the **Analyze** page in this dashboard to run analysis interactively.",
        ),
        (
            "What are deep static adapters?",
            "Adapters powered by real analysis tools instead of regex. "
            "Use `--enable-adapters` to activate Lizard (complexity), "
            "Flawfinder (security), and CCLS/libclang (dead code, call graphs, function metrics).",
        ),
    ]
    for q, a in FAQS:
        with st.expander(q):
            st.markdown(a)

    st.divider()
    net_ip = st_tools.get_local_ip()
    st.markdown(
        f"**Dashboard access:**  \n"
        f"This machine: [http://localhost:8502](http://localhost:8502)  \n"
        f"Network: [http://{net_ip}:8502](http://{net_ip}:8502)  \n\n"
        f"**Contact:** sendpavanr@gmail.com  \n"
        f"**Model:** `{STREAMLIT_MODEL}`"
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    st_tools.app_css()

    # Header logo
    if os.path.isfile(LOGO_MAIN):
        st.image(LOGO_MAIN, width=480)

    # Sidebar navigation (returns "Workflow", "Chat", or "About")
    page = st_tools.sidebar(LOGO_SIDEBAR)

    if page == "Chat":
        page_chat()
    elif page == "About":
        page_about()
    else:
        # Workflow tabs — the seven main workflow stages
        (tab_analyze, tab_pipeline, tab_review, tab_fixqa,
         tab_audit, tab_constraints, tab_telemetry) = st.tabs([
            "📊 Analyze", "⚙️ Pipeline", "📝 Review", "🔧 Fix & QA",
            "📋 Audit", "📐 Constraints", "📈 Telemetry",
        ])
        with tab_analyze:
            page_analyze()
        with tab_pipeline:
            page_pipeline()
        with tab_review:
            page_review()
        with tab_fixqa:
            page_fixer_qa()
        with tab_audit:
            page_fixer_audit()
        with tab_constraints:
            page_constraints_generator()
        with tab_telemetry:
            page_telemetry()


if __name__ == "__main__":
    main()
